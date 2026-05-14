"""
Test per:
  - Navigazione settimanale (/settimane/{hotel_code}, /settimane/gruppo)
  - Logica confronto anno precedente (stessa settimana commerciale sab–ven)
  - Logica confronto settimana precedente
  - Export Excel e CSV con dati reali
"""

import os
import sys
import io
from datetime import date, timedelta

import pytest

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from sqlalchemy import create_engine, text
from sqlalchemy.orm import sessionmaker
from starlette.testclient import TestClient

from app.database import Base, get_db
from app.main import app
from app.models.revenue import DailyRevenue, ImportSession  # noqa: F401
from app.services.weekly_aggregator import settimana_di

UPLOADS_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..", "uploads"))
TEST_DB_URL = "postgresql://ginoscola@localhost:5432/revenue_master_test"

# Salta i test che richiedono i file CSV reali se non disponibili
_FILE_CSV_DISPONIBILI = os.path.isfile(os.path.join(UPLOADS_DIR, "PlanningForecast-CLB1.csv"))
richiede_csv = pytest.mark.skipif(
    not _FILE_CSV_DISPONIBILI,
    reason="File CSV reali non trovati in uploads/ — copiare PlanningForecast-*.csv prima di eseguire",
)


# ---------------------------------------------------------------------------
# Fixture DB di test
# ---------------------------------------------------------------------------

@pytest.fixture(scope="module")
def test_engine():
    engine = create_engine(TEST_DB_URL)
    Base.metadata.create_all(engine)
    yield engine
    Base.metadata.drop_all(engine)
    engine.dispose()


@pytest.fixture(scope="module")
def TestSession(test_engine):
    return sessionmaker(bind=test_engine)


@pytest.fixture(scope="module")
def client(test_engine, TestSession):
    def override_get_db():
        db = TestSession()
        try:
            yield db
        finally:
            db.close()
    app.dependency_overrides[get_db] = override_get_db
    with TestClient(app) as c:
        yield c
    app.dependency_overrides.clear()


@pytest.fixture(autouse=True)
def pulisci_tabelle(test_engine):
    """Svuota le tabelle e importa i file reali prima di ogni test."""
    with test_engine.connect() as conn:
        conn.execute(text(
            "TRUNCATE daily_revenue, hotel_seasons, hotels, imports RESTART IDENTITY CASCADE"
        ))
        conn.execute(text("""
            INSERT INTO hotels (code, name, default_rooms) VALUES
            ('CLB', 'Club Hotel', 45),
            ('DPH', 'Hotel Du Parc', 43),
            ('INT', 'Hotel International', 45)
        """))
        conn.commit()


def _percorso(nome):
    return os.path.join(UPLOADS_DIR, nome)


def _importa_hotel(client, hotel_code):
    """Importa la coppia di file reali per un hotel nel DB di test."""
    with (
        open(_percorso(f"PlanningForecast-{hotel_code}1.csv"), "rb") as f1,
        open(_percorso(f"PlanningForecast-{hotel_code}2.csv"), "rb") as f2,
    ):
        resp = client.post(
            f"/upload/coppia/{hotel_code}",
            files={"file1": (f"PlanningForecast-{hotel_code}1.csv", f1), "file2": (f"PlanningForecast-{hotel_code}2.csv", f2)},
        )
    assert resp.status_code == 200
    return resp.json()


# ---------------------------------------------------------------------------
# Test: navigazione settimanale — hotel
# ---------------------------------------------------------------------------

class TestListaSettimaneHotel:
    def test_lista_vuota_senza_dati(self, client):
        resp = client.get("/settimane/CLB")
        assert resp.status_code == 200
        assert resp.json()["settimane"] == []

    @richiede_csv
    def test_lista_non_vuota_dopo_import(self, client):
        _importa_hotel(client, "CLB")
        resp = client.get("/settimane/CLB")
        j = resp.json()
        assert len(j["settimane"]) > 0

    @richiede_csv
    def test_prima_settimana_e_la_piu_recente(self, client):
        _importa_hotel(client, "CLB")
        j = client.get("/settimane/CLB").json()
        settimane = j["settimane"]
        # ordinate dalla più recente: la prima deve avere week_start > dell'ultima
        assert settimane[0]["week_start"] > settimane[-1]["week_start"]

    @richiede_csv
    def test_ultima_settimana_e_la_piu_antica(self, client):
        _importa_hotel(client, "CLB")
        j = client.get("/settimane/CLB").json()
        settimane = j["settimane"]
        date_ws = [date.fromisoformat(s["week_start"]) for s in settimane]
        assert date_ws == sorted(date_ws, reverse=True)

    @richiede_csv
    def test_week_start_e_sempre_sabato(self, client):
        _importa_hotel(client, "CLB")
        j = client.get("/settimane/CLB").json()
        for s in j["settimane"]:
            ws = date.fromisoformat(s["week_start"])
            assert ws.weekday() == 5, f"{ws} non è sabato"

    @richiede_csv
    def test_week_end_e_sempre_venerdi(self, client):
        _importa_hotel(client, "CLB")
        j = client.get("/settimane/CLB").json()
        for s in j["settimane"]:
            we = date.fromisoformat(s["week_end"])
            assert we.weekday() == 4, f"{we} non è venerdì"

    @richiede_csv
    def test_label_formato_italiano(self, client):
        _importa_hotel(client, "CLB")
        j = client.get("/settimane/CLB").json()
        # La label deve contenere nomi mese in italiano
        mesi = ["gen","feb","mar","apr","mag","giu","lug","ago","set","ott","nov","dic"]
        for s in j["settimane"]:
            assert any(m in s["label"] for m in mesi), f"Nessun mese italiano in '{s['label']}'"

    @richiede_csv
    def test_numero_settimane_clb(self, client):
        """CLB: 113 giorni → 17 settimane commerciali."""
        _importa_hotel(client, "CLB")
        j = client.get("/settimane/CLB").json()
        assert len(j["settimane"]) == 17

    @richiede_csv
    def test_snapshot_date_presente(self, client):
        _importa_hotel(client, "CLB")
        j = client.get("/settimane/CLB").json()
        # Almeno una settimana deve avere snapshot_date
        assert any(s["snapshot_date"] is not None for s in j["settimane"])


# ---------------------------------------------------------------------------
# Test: navigazione settimanale — gruppo
# ---------------------------------------------------------------------------

class TestListaSettimaneGruppo:
    @richiede_csv
    def test_gruppo_unisce_settimane_di_tutti_gli_hotel(self, client):
        for code in ["CLB", "DPH", "INT"]:
            _importa_hotel(client, code)
        j_clb   = client.get("/settimane/CLB").json()
        j_dph   = client.get("/settimane/DPH").json()
        j_gruppo = client.get("/settimane/gruppo").json()

        ws_clb   = {s["week_start"] for s in j_clb["settimane"]}
        ws_dph   = {s["week_start"] for s in j_dph["settimane"]}
        ws_gruppo = {s["week_start"] for s in j_gruppo["settimane"]}

        # Il gruppo deve contenere almeno tutte le settimane dei singoli hotel
        assert ws_clb.issubset(ws_gruppo)
        assert ws_dph.issubset(ws_gruppo)

    @richiede_csv
    def test_gruppo_piu_settimane_di_ogni_hotel_singolo(self, client):
        """DPH apre il 01/05, CLB il 01/06: il gruppo ha più settimane di CLB."""
        for code in ["CLB", "DPH", "INT"]:
            _importa_hotel(client, code)
        n_clb   = len(client.get("/settimane/CLB").json()["settimane"])
        n_gruppo = len(client.get("/settimane/gruppo").json()["settimane"])
        assert n_gruppo > n_clb


# ---------------------------------------------------------------------------
# Test: logica confronto settimana precedente
# ---------------------------------------------------------------------------

class TestConfrontoSettimanaPrecedente:
    def test_settimana_prec_e_sempre_7_giorni_prima(self):
        """La settimana precedente di qualsiasi sabato è sabato -7 giorni."""
        sabato = date(2026, 6, 6)
        sett_prec_start = sabato - timedelta(days=7)
        sett_prec_end   = sett_prec_start + timedelta(days=6)
        assert sett_prec_start.weekday() == 5  # sabato
        assert sett_prec_end.weekday() == 4    # venerdì
        assert sett_prec_start == date(2026, 5, 30)

    @richiede_csv
    def test_confronto_sett_prec_endpoint(self, client):
        """L'endpoint dashboard con periodo di una sola settimana funziona."""
        _importa_hotel(client, "CLB")
        j = client.get("/settimane/CLB").json()
        settimane = j["settimane"]
        if len(settimane) < 2:
            pytest.skip("Meno di 2 settimane disponibili")

        # Prima (più recente)
        ws1 = settimane[0]["week_start"]
        we1 = settimane[0]["week_end"]
        resp1 = client.get(f"/dashboard/hotel/CLB?da={ws1}&a={we1}")
        assert resp1.status_code == 200

        # Settimana precedente (seconda in lista)
        ws2 = settimane[1]["week_start"]
        we2 = settimane[1]["week_end"]
        resp2 = client.get(f"/dashboard/hotel/CLB?da={ws2}&a={we2}")
        assert resp2.status_code == 200

        # I KPI devono essere diversi (settimane diverse)
        k1 = resp1.json()["kpi_periodo"]
        k2 = resp2.json()["kpi_periodo"]
        # Almeno rooms_sold o revenue devono differire tra le due settimane
        assert k1["rooms_sold"] != k2["rooms_sold"] or k1["adr"] != k2["adr"]


# ---------------------------------------------------------------------------
# Test: logica confronto anno precedente
# ---------------------------------------------------------------------------

class TestConfrontoAnnoPrecedente:
    def test_meno_364_giorni_e_stesso_giorno_settimana(self):
        """Sottraendo 364 giorni (52 settimane) si ottiene sempre lo stesso giorno della settimana."""
        for sabato in [
            date(2026, 6, 6),
            date(2026, 7, 4),
            date(2026, 8, 1),
            date(2026, 9, 12),
        ]:
            anno_prec = sabato - timedelta(days=364)
            assert anno_prec.weekday() == 5, (
                f"{sabato} - 364 = {anno_prec} non è sabato (weekday={anno_prec.weekday()})"
            )

    def test_settimana_anno_prec_e_commerciale(self):
        """La settimana dell'anno precedente (−364) è una settimana commerciale valida."""
        sabato_2026 = date(2026, 6, 6)
        sabato_2025 = sabato_2026 - timedelta(days=364)
        venerdi_2025 = sabato_2025 + timedelta(days=6)
        assert sabato_2025 == settimana_di(sabato_2025)
        assert venerdi_2025.weekday() == 4

    @richiede_csv
    def test_confronto_anno_prec_dati_non_disponibili(self, client):
        """Se i dati dell'anno precedente non esistono, l'endpoint restituisce 404."""
        _importa_hotel(client, "CLB")
        j = client.get("/settimane/CLB").json()
        ws = j["settimane"][0]["week_start"]
        we = j["settimane"][0]["week_end"]

        # Anno precedente: -364 giorni (nessun dato nel test DB)
        ws_prec = (date.fromisoformat(ws) - timedelta(days=364)).isoformat()
        we_prec = (date.fromisoformat(we) - timedelta(days=364)).isoformat()
        resp = client.get(f"/dashboard/hotel/CLB?da={ws_prec}&a={we_prec}")
        assert resp.status_code == 404


# ---------------------------------------------------------------------------
# Test: export Excel
# ---------------------------------------------------------------------------

class TestExportExcel:
    @richiede_csv
    def test_export_xlsx_hotel_settimanale(self, client):
        _importa_hotel(client, "CLB")
        resp = client.get("/export/hotel/CLB/settimanale?formato=xlsx")
        assert resp.status_code == 200
        assert "spreadsheetml" in resp.headers["content-type"]
        assert "CLB_settimanale.xlsx" in resp.headers.get("content-disposition", "")

    @richiede_csv
    def test_export_xlsx_contiene_righe_reali(self, client):
        _importa_hotel(client, "CLB")
        resp = client.get("/export/hotel/CLB/settimanale?formato=xlsx")
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(resp.content))
        ws = wb.active
        # 17 settimane + 1 intestazione + 1 totale stagione
        assert ws.max_row == 19

    @richiede_csv
    def test_export_xlsx_intestazione_bold(self, client):
        _importa_hotel(client, "CLB")
        resp = client.get("/export/hotel/CLB/settimanale?formato=xlsx")
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(resp.content))
        ws = wb.active
        assert ws.cell(row=1, column=1).font.bold is True

    @richiede_csv
    def test_export_xlsx_giornaliero(self, client):
        _importa_hotel(client, "CLB")
        resp = client.get("/export/hotel/CLB/giornaliero?formato=xlsx")
        assert resp.status_code == 200
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(resp.content))
        ws = wb.active
        # 113 giorni CLB + 1 intestazione
        assert ws.max_row == 114

    @richiede_csv
    def test_export_xlsx_gruppo(self, client):
        for code in ["CLB", "DPH", "INT"]:
            _importa_hotel(client, code)
        resp = client.get("/export/gruppo?formato=xlsx")
        assert resp.status_code == 200
        assert "spreadsheetml" in resp.headers["content-type"]

    @richiede_csv
    def test_export_xlsx_hotel_sett_16_colonne(self, client):
        """Il foglio settimanale hotel deve avere esattamente 16 colonne."""
        _importa_hotel(client, "CLB")
        resp = client.get("/export/hotel/CLB/settimanale?formato=xlsx")
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(resp.content))
        ws = wb.active
        assert ws.max_column == 16
        # Verifica prima e ultima intestazione
        assert ws.cell(row=1, column=1).value == "Settimana"
        assert "Inc." in str(ws.cell(row=1, column=16).value)

    @richiede_csv
    def test_export_xlsx_hotel_sett_riga_totale_bold(self, client):
        """L'ultima riga dell'export settimanale deve essere la riga totale in grassetto."""
        _importa_hotel(client, "CLB")
        resp = client.get("/export/hotel/CLB/settimanale?formato=xlsx")
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(resp.content))
        ws = wb.active
        ultima = ws.max_row
        assert ws.cell(row=ultima, column=1).font.bold is True
        assert "TOTALE" in str(ws.cell(row=ultima, column=1).value).upper()

    @richiede_csv
    def test_export_xlsx_gruppo_due_fogli(self, client):
        """L'export gruppo Excel deve avere 2 fogli: aggregati settimanali e dettaglio hotel."""
        for code in ["CLB", "DPH", "INT"]:
            _importa_hotel(client, code)
        resp = client.get("/export/gruppo?formato=xlsx")
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(resp.content))
        assert len(wb.sheetnames) == 2
        assert "Aggregati" in wb.sheetnames[0]
        assert "Dettaglio" in wb.sheetnames[1]


# ---------------------------------------------------------------------------
# Test: export CSV
# ---------------------------------------------------------------------------

class TestExportCSV:
    @richiede_csv
    def test_export_csv_hotel_settimanale(self, client):
        _importa_hotel(client, "CLB")
        resp = client.get("/export/hotel/CLB/settimanale?formato=csv")
        assert resp.status_code == 200
        assert "csv" in resp.headers["content-type"]

    @richiede_csv
    def test_export_csv_ha_intestazione_e_righe(self, client):
        _importa_hotel(client, "CLB")
        resp = client.get("/export/hotel/CLB/settimanale?formato=csv")
        import csv as csv_mod
        contenuto = resp.content.decode("utf-8-sig")
        reader = list(csv_mod.reader(contenuto.splitlines()))
        # 1 intestazione + 17 settimane + 1 totale stagione
        assert len(reader) == 19
        assert "Settimana" in reader[0][0]

    @richiede_csv
    def test_export_csv_hotel_sett_16_colonne(self, client):
        """Il CSV settimanale hotel deve avere 16 colonne e includere Inc. Rooms/F&B/Extra."""
        _importa_hotel(client, "CLB")
        resp = client.get("/export/hotel/CLB/settimanale?formato=csv")
        import csv as csv_mod
        contenuto = resp.content.decode("utf-8-sig")
        reader = list(csv_mod.reader(contenuto.splitlines()))
        intestazione = reader[0]
        assert len(intestazione) == 16
        assert any("Inc." in col for col in intestazione)
        assert any("TRevPAR" in col for col in intestazione)
        assert any("RMC" in col for col in intestazione)

    @richiede_csv
    def test_export_csv_giornaliero_ha_113_righe_clb(self, client):
        _importa_hotel(client, "CLB")
        resp = client.get("/export/hotel/CLB/giornaliero?formato=csv")
        import csv as csv_mod
        contenuto = resp.content.decode("utf-8-sig")
        reader = list(csv_mod.reader(contenuto.splitlines()))
        assert len(reader) == 114  # 113 dati + 1 intestazione

    @richiede_csv
    def test_export_csv_con_filtro_periodo(self, client):
        _importa_hotel(client, "CLB")
        # Solo prima settimana commerciale CLB: 30/05–05/06/2026
        resp = client.get("/export/hotel/CLB/giornaliero?da=2026-05-30&a=2026-06-05&formato=csv")
        import csv as csv_mod
        contenuto = resp.content.decode("utf-8-sig")
        reader = list(csv_mod.reader(contenuto.splitlines()))
        assert len(reader) == 8  # 7 giorni + 1 intestazione

    @richiede_csv
    def test_export_pdf_hotel_settimanale(self, client):
        _importa_hotel(client, "CLB")
        resp = client.get("/export/hotel/CLB/settimanale?formato=pdf")
        assert resp.status_code == 200
        assert resp.headers["content-type"] == "application/pdf"
        assert resp.content[:4] == b"%PDF"


# ---------------------------------------------------------------------------
# Test: kpi_stagione (modalità snapshot — intera stagione)
# ---------------------------------------------------------------------------

class TestKpiStagioneSnapshot:
    @richiede_csv
    def test_kpi_stagione_presente_nella_risposta(self, client):
        """La risposta in modalità snapshot include il campo kpi_stagione."""
        _importa_hotel(client, "CLB")
        snap = client.get("/snapshots/CLB").json()["snapshots"]
        assert len(snap) > 0
        sd = snap[0]["snapshot_date"]
        resp = client.get(f"/dashboard/hotel/CLB?snapshot={sd}")
        assert resp.status_code == 200
        assert "kpi_stagione" in resp.json()

    @richiede_csv
    def test_kpi_stagione_copre_tutta_la_stagione(self, client):
        """kpi_stagione.rooms_sold deve essere pari alla somma di tutti i giorni della snapshot."""
        _importa_hotel(client, "CLB")
        snap = client.get("/snapshots/CLB").json()["snapshots"]
        sd = snap[0]["snapshot_date"]
        resp = client.get(f"/dashboard/hotel/CLB?snapshot={sd}").json()

        # Somma rooms_sold da tutti i giorni della stagione
        totale_giorni = sum(g["rooms_sold"] for g in resp["giorni"])
        assert resp["kpi_stagione"]["rooms_sold"] == totale_giorni

    @richiede_csv
    def test_kpi_stagione_diverso_da_singola_settimana(self, client):
        """kpi_stagione deve avere rooms_sold > della singola settimana di riferimento."""
        _importa_hotel(client, "CLB")
        snap = client.get("/snapshots/CLB").json()["snapshots"]
        sd = snap[0]["snapshot_date"]
        resp = client.get(f"/dashboard/hotel/CLB?snapshot={sd}").json()

        kpi_stagione = resp["kpi_stagione"]["rooms_sold"]
        ref_start = resp["settimana_ref_start"]
        ref_end = resp["settimana_ref_end"]

        if ref_start and ref_end:
            rooms_ref = sum(
                g["rooms_sold"] for g in resp["giorni"]
                if ref_start <= g["data"] <= ref_end
            )
            # La stagione intera deve avere più camere della sola settimana di rif.
            assert kpi_stagione >= rooms_ref

    @richiede_csv
    def test_snapshots_endpoint_restituisce_lista(self, client):
        """GET /snapshots/{hotel_code} restituisce una lista ordinata dalla più recente."""
        _importa_hotel(client, "CLB")
        resp = client.get("/snapshots/CLB")
        assert resp.status_code == 200
        snaps = resp.json()["snapshots"]
        assert len(snaps) > 0
        assert "snapshot_date" in snaps[0]
        assert "label" in snaps[0]
        # Ordine discendente
        if len(snaps) > 1:
            dates = [s["snapshot_date"] for s in snaps]
            assert dates == sorted(dates, reverse=True)

    @richiede_csv
    def test_export_con_snapshot_filtra_correttamente(self, client):
        """L'export con ?snapshot= deve contenere solo i dati di quella snapshot."""
        _importa_hotel(client, "CLB")
        snap = client.get("/snapshots/CLB").json()["snapshots"]
        sd = snap[0]["snapshot_date"]
        resp = client.get(f"/export/hotel/CLB/giornaliero?snapshot={sd}&formato=csv")
        assert resp.status_code == 200
        import csv as csv_mod
        contenuto = resp.content.decode("utf-8-sig")
        reader = list(csv_mod.reader(contenuto.splitlines()))
        # 113 giorni CLB + 1 intestazione
        assert len(reader) == 114
