"""Router modulo Corrispettivi v4 — tabella analitica unificata.

Endpoint:
  POST   /corrispettivi/import                → upload file Excel (on_conflict: salta|aggiorna)
  GET    /corrispettivi/import/storico        → storico sessioni import
  DELETE /corrispettivi/import/{id}           → elimina import e documenti collegati

  GET    /corrispettivi/documenti             → lista unificata con filtri e paginazione
  GET    /corrispettivi/scontrini             → alias documenti tipo=scontrino
  GET    /corrispettivi/fatture               → alias documenti tipo=fattura
  PUT    /corrispettivi/documenti/{id}        → correzione manuale unificata
  PUT    /corrispettivi/scontrini/{id}        → alias → PUT /documenti/{id}
  PUT    /corrispettivi/fatture/{id}          → alias → PUT /documenti/{id}

  POST   /corrispettivi/manuali               → inserimento MMS/BON
  PUT    /corrispettivi/manuali/{id}          → modifica MMS/BON
  GET    /corrispettivi/manuali               → lista con filtri

  GET    /corrispettivi/report/giornaliero    → aggregato per giorno e struttura
  GET    /corrispettivi/report/mensile        → aggregato per mese e struttura
  GET    /corrispettivi/check                 → totali per struttura

  GET    /corrispettivi/admin/test-stats      → conteggio record is_test
  DELETE /corrispettivi/admin/test-data       → cancella tutti i record is_test

  POST   /corrispettivi/rt-chiusure           → upsert chiusura RT giornaliera (admin)
  GET    /corrispettivi/rt-chiusure           → lista mese con delta vs PMS
  DELETE /corrispettivi/rt-chiusure/{id}      → elimina chiusura RT (admin)
"""

import os
import tempfile
from datetime import date, datetime
from decimal import Decimal
from typing import List, Optional, Set, Tuple

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from sqlalchemy import func, text, update
from sqlalchemy.dialects.postgresql import insert as pg_insert
from sqlalchemy.orm import Session

from app.database import get_db
from app.auth import richiedi_admin, richiedi_utente_attivo
from app.models.corrispettivi import (
    CorrispettiviDocumento,
    CorrispettiviImport,
    CorrispettiviManuale,
    RtChiusura,
)
from app.services.corrispettivi_excel_parser import _determina_categoria, parse_excel

router = APIRouter(prefix="/corrispettivi", tags=["corrispettivi"])

STRUTTURE_HOTEL: List[str] = ['DPH', 'CLB', 'INT']
STRUTTURE_MANUALI: List[str] = ['MMS', 'BON']
STRUTTURE_ORDINE: List[str] = STRUTTURE_HOTEL + STRUTTURE_MANUALI

NOME_STRUTTURA = {
    'DPH': 'Hotel Du Parc',
    'CLB': 'Club Hotel',
    'INT': 'Hotel International',
    'MMS': 'Maremosso',
    'BON': 'Buona Onda',
}

CATEGORIE: List[str] = ['arrangiamenti', 'tassa_soggiorno', 'penali', 'shop', 'altro']

IVA_MANUALI_PCT = 10.0


# ── Helpers ───────────────────────────────────────────────────────────────────

def _to_float(v) -> float:
    if v is None:
        return 0.0
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def _d(v) -> Decimal:
    try:
        f = float(v or 0)
        if f != f:  # NaN
            f = 0.0
        return Decimal(str(round(f, 2)))
    except (TypeError, ValueError):
        return Decimal('0')


def _fmt_import(imp: CorrispettiviImport) -> dict:
    return {
        'id': imp.id,
        'nome_file': imp.nome_file,
        'data_da': imp.data_da.isoformat() if imp.data_da else None,
        'data_a': imp.data_a.isoformat() if imp.data_a else None,
        'tipo_import': imp.tipo_import,
        'strutture_presenti': imp.strutture_presenti or [],
        'n_scontrini': imp.n_scontrini,
        'n_fatture': imp.n_fatture,
        'n_esclusi': imp.n_esclusi,
        'is_test': imp.is_test,
        'created_at': imp.created_at.isoformat() if imp.created_at else None,
    }


def _disaggrega_imponibile(d: CorrispettiviDocumento) -> dict:
    """Calcola imponibile e IVA disaggregati per arrangiamenti e tassa soggiorno.

    Restituisce chiavi aggiuntive da includere in _fmt_documento.
    Significativo solo per categoria='arrangiamenti'; per le altre categorie
    i campi sono None per evitare confusione.
    """
    if d.categoria != 'arrangiamenti':
        return {
            'imponibile_arr': None, 'iva_arr': None,
            'imponibile_ts': None,  'iva_ts': None,
        }

    totale = _to_float(d.totale_lordo)
    ts_raw = None if d.tassa_soggiorno is None else _to_float(d.tassa_soggiorno)
    sum_iva = _to_float(d.iva)

    if ts_raw is not None:
        # Formato esteso: valore esatto
        lordo_ts  = round(ts_raw, 2)
        lordo_arr = round(max(0.0, totale - lordo_ts), 2)
    elif sum_iva > 0:
        # Formato base: inferenza da IVA (imp_arr = iva*10, lordo_arr = iva*11)
        lordo_arr = round(sum_iva * 11, 2)
        lordo_ts  = round(max(0.0, totale - lordo_arr), 2)
    else:
        # Nessuna tassa soggiorno rilevabile
        lordo_arr = totale
        lordo_ts  = 0.0

    imp_arr = round(lordo_arr / 1.10, 2) if lordo_arr else 0.0
    iva_arr  = round(lordo_arr - imp_arr, 2)
    imp_ts   = round(lordo_ts, 2)   # aliquota 0% → imponibile = lordo
    iva_ts   = 0.0

    return {
        'imponibile_arr': imp_arr,
        'iva_arr':        iva_arr,
        'imponibile_ts':  imp_ts,
        'iva_ts':         iva_ts,
    }


def _fmt_documento(d: CorrispettiviDocumento) -> dict:
    return {
        'id': d.id,
        'import_id': d.import_id,
        'tipo': d.tipo,
        'data_documento': d.data_documento.isoformat() if d.data_documento else None,
        'numero': d.numero,
        'suffisso': d.suffisso,
        'struttura_code': d.struttura_code,
        'struttura_nome': NOME_STRUTTURA.get(d.struttura_code, d.struttura_code),
        'intestazione': d.intestazione,
        'camera': d.camera,
        'totale_lordo': _to_float(d.totale_lordo),
        'incassato': _to_float(d.incassato),
        'deposito': _to_float(d.deposito),
        'sospeso': _to_float(d.sospeso),
        'abbuono': _to_float(d.abbuono),
        'imponibile': _to_float(d.imponibile),
        'iva': _to_float(d.iva),
        'aliquota_pct': _to_float(d.aliquota_pct),
        'tassa_soggiorno': None if d.tassa_soggiorno is None else _to_float(d.tassa_soggiorno),
        'categoria': d.categoria,
        'codice_prenotazione': d.codice_prenotazione,
        'tipo_pagamento': d.tipo_pagamento,
        'categoria_pagamento': d.categoria_pagamento,
        **_disaggrega_imponibile(d),
        'conto_anticipato': d.conto_anticipato,
        'acconto': d.acconto,
        'annullato': d.annullato,
        'ospiti': d.ospiti,
        'note': d.note,
        'motivo_esclusione': d.motivo_esclusione,
        'modificato_manualmente': d.modificato_manualmente,
        'modifica_note': d.modifica_note,
        'is_test': d.is_test,
        # Campi formato esteso Welcome PMS
        'sigla': d.sigla,
        'numero_scontrino': d.numero_scontrino,
        'arrivo': d.arrivo.isoformat() if d.arrivo else None,
        'partenza': d.partenza.isoformat() if d.partenza else None,
        'ubicazione_istat': d.ubicazione_istat,
        'voucher': d.voucher,
        'nome_file_pms': d.nome_file_pms,
        'stato_fe': d.stato_fe,
        'modalita': d.modalita,
        'importo_bollo': None if d.importo_bollo is None else _to_float(d.importo_bollo),
        'tipo_documento_fe': d.tipo_documento_fe,
        'numero_documento_fe': d.numero_documento_fe,
        'nazione': d.nazione,
        'ora_stampa': d.ora_stampa,
        'contabilizzato_mexal': d.contabilizzato_mexal,
        'causale_cancellazione': d.causale_cancellazione,
        'maschera_conto': d.maschera_conto,
        'data_creazione_doc': d.data_creazione_doc.isoformat() if d.data_creazione_doc else None,
        'utente_creazione': d.utente_creazione,
    }


def _fmt_manuale(m: CorrispettiviManuale) -> dict:
    lordo = _to_float(m.arrangiamenti_lordo)
    imponibile = round(lordo / 1.10, 2)
    iva = round(lordo - imponibile, 2)
    return {
        'id': m.id,
        'data_giorno': m.data_giorno.isoformat() if m.data_giorno else None,
        'struttura_code': m.struttura_code,
        'struttura_nome': NOME_STRUTTURA.get(m.struttura_code, m.struttura_code),
        'arrangiamenti_lordo': lordo,
        'arrangiamenti_imponibile': imponibile,
        'arrangiamenti_iva': iva,
        'note': m.note,
        'is_test': m.is_test,
        'updated_at': m.updated_at.isoformat() if m.updated_at else None,
    }


# ── POST /corrispettivi/import ────────────────────────────────────────────────

@router.post("/import")
def importa_excel(
    file: UploadFile = File(...),
    is_test: bool = Query(False),
    on_conflict: str = Query('salta', description="'salta' (DO NOTHING) o 'aggiorna' (aggiorna se non modificato manualmente)"),
    db: Session = Depends(get_db),
    utente=Depends(richiedi_admin),
):
    """Carica un file Excel listaConti.xlsx e importa tutti i documenti."""
    if not file.filename.lower().endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Solo file Excel (.xlsx) accettati")
    if on_conflict not in ('salta', 'aggiorna'):
        raise HTTPException(status_code=400, detail="on_conflict deve essere 'salta' o 'aggiorna'")

    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
        tmp.write(file.file.read())
        tmp_path = tmp.name

    try:
        risultato = parse_excel(tmp_path)
    except Exception as exc:
        os.unlink(tmp_path)
        raise HTTPException(status_code=500, detail=f"Errore parser: {exc}")
    else:
        os.unlink(tmp_path)

    if not risultato.scontrini and not risultato.fatture:
        raise HTTPException(
            status_code=422,
            detail="Nessun documento SC/SCA/F trovato nel file. "
                   "Verifica che sia un file listaConti.xlsx esportato da Welcome PMS.",
        )

    # Tutti i documenti da salvare (inclusi esclusi per audit)
    tutti = risultato.documenti

    # Pre-check: quali chiavi esistono già in DB?
    chiavi_tutti = [
        (d.struttura_code, d.data_documento, d.numero, d.suffisso)
        for d in tutti
    ]
    strutture_set = {d.struttura_code for d in tutti}
    date_set = {d.data_documento for d in tutti}

    if chiavi_tutti:
        esistenti_q = db.query(
            CorrispettiviDocumento.struttura_code,
            CorrispettiviDocumento.data_documento,
            CorrispettiviDocumento.numero,
            CorrispettiviDocumento.suffisso,
            CorrispettiviDocumento.modificato_manualmente,
        ).filter(
            CorrispettiviDocumento.struttura_code.in_(strutture_set),
            CorrispettiviDocumento.data_documento.in_(date_set),
        ).all()

        chiavi_esistenti: dict = {
            (r.struttura_code, r.data_documento, r.numero, r.suffisso): r.modificato_manualmente
            for r in esistenti_q
        }
    else:
        chiavi_esistenti = {}

    nuovi = []
    aggiornabili = []
    protetti = []

    for d in tutti:
        chiave = (d.struttura_code, d.data_documento, d.numero, d.suffisso)
        if chiave not in chiavi_esistenti:
            nuovi.append(d)
        elif chiavi_esistenti[chiave]:   # modificato_manualmente=True
            protetti.append(d)
        else:
            aggiornabili.append(d)

    # Sessione import
    imp = CorrispettiviImport(
        nome_file=file.filename,
        data_da=risultato.data_da,
        data_a=risultato.data_a,
        tipo_import='excel',
        strutture_presenti=sorted(risultato.strutture_trovate),
        n_scontrini=len(risultato.scontrini),
        n_fatture=len(risultato.fatture),
        n_esclusi=len(risultato.esclusi),
        is_test=is_test,
        imported_by=utente.id,
    )
    db.add(imp)
    db.flush()   # ottiene imp.id

    def _campi_estesi(d) -> dict:
        """Campi formato esteso Welcome PMS (None se non disponibili)."""
        return dict(
            tassa_soggiorno=_d(d.tassa_soggiorno) if d.tassa_soggiorno is not None else None,
            sigla=d.sigla,
            numero_scontrino=d.numero_scontrino,
            arrivo=d.arrivo,
            partenza=d.partenza,
            ubicazione_istat=d.ubicazione_istat,
            voucher=d.voucher,
            nome_file_pms=d.nome_file_pms,
            stato_fe=d.stato_fe,
            modalita=d.modalita,
            importo_bollo=_d(d.importo_bollo) if d.importo_bollo is not None else None,
            tipo_documento_fe=d.tipo_documento_fe,
            numero_documento_fe=d.numero_documento_fe,
            nazione=d.nazione,
            ora_stampa=d.ora_stampa,
            contabilizzato_mexal=d.contabilizzato_mexal,
            causale_cancellazione=d.causale_cancellazione,
            maschera_conto=d.maschera_conto,
            data_creazione_doc=d.data_creazione_doc,
            utente_creazione=d.utente_creazione,
        )

    def _valori_doc(d, import_id: int) -> dict:
        return dict(
            import_id=import_id,
            data_documento=d.data_documento,
            numero=d.numero,
            suffisso=d.suffisso,
            tipo=d.tipo,
            struttura_code=d.struttura_code,
            intestazione=d.intestazione or '',
            camera=d.camera or '',
            totale_lordo=_d(d.totale_lordo),
            incassato=_d(d.incassato),
            deposito=_d(d.deposito),
            sospeso=_d(d.sospeso),
            abbuono=_d(d.abbuono),
            imponibile=_d(d.imponibile),
            iva=_d(d.iva),
            aliquota_pct=_d(d.aliquota_pct),
            categoria=d.categoria,
            codice_prenotazione=d.codice_prenotazione or '',
            tipo_pagamento=d.tipo_pagamento or '',
            conto_anticipato=d.conto_anticipato,
            acconto=d.acconto,
            annullato=d.annullato,
            ospiti=d.ospiti or '',
            note=d.note or '',
            motivo_esclusione=d.motivo_esclusione,
            modificato_manualmente=False,
            is_test=is_test,
            **_campi_estesi(d),
        )

    try:
        # Inserisce nuovi (DO NOTHING come safety net per race condition)
        for d in nuovi:
            v = _valori_doc(d, imp.id)
            stmt = pg_insert(CorrispettiviDocumento).values(**v)
            stmt = stmt.on_conflict_do_nothing(constraint='uq_documento')
            db.execute(stmt)

        # Se "aggiorna": aggiorna quelli esistenti non protetti (import_id rimane invariato — Option A)
        n_aggiornati = 0
        if on_conflict == 'aggiorna' and aggiornabili:
            for d in aggiornabili:
                db.execute(
                    update(CorrispettiviDocumento)
                    .where(
                        CorrispettiviDocumento.struttura_code == d.struttura_code,
                        CorrispettiviDocumento.data_documento == d.data_documento,
                        CorrispettiviDocumento.numero == d.numero,
                        CorrispettiviDocumento.suffisso == d.suffisso,
                        CorrispettiviDocumento.modificato_manualmente == False,
                    )
                    .values(
                        tipo=d.tipo,
                        intestazione=d.intestazione or '',
                        camera=d.camera or '',
                        totale_lordo=_d(d.totale_lordo),
                        incassato=_d(d.incassato),
                        deposito=_d(d.deposito),
                        sospeso=_d(d.sospeso),
                        abbuono=_d(d.abbuono),
                        imponibile=_d(d.imponibile),
                        iva=_d(d.iva),
                        aliquota_pct=_d(d.aliquota_pct),
                        categoria=d.categoria,
                        codice_prenotazione=d.codice_prenotazione or '',
                        tipo_pagamento=d.tipo_pagamento or '',
                        conto_anticipato=d.conto_anticipato,
                        acconto=d.acconto,
                        annullato=d.annullato,
                        ospiti=d.ospiti or '',
                        note=d.note or '',
                        motivo_esclusione=d.motivo_esclusione,
                        is_test=is_test,
                        # import_id NON aggiornato (Option A)
                        **_campi_estesi(d),
                    )
                )
            n_aggiornati = len(aggiornabili)

        db.commit()
    except Exception as exc:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Errore DB: {exc}")

    return {
        'id': imp.id,
        'n_inseriti': len(nuovi),
        'n_aggiornati': n_aggiornati,
        'n_saltati': len(aggiornabili) if on_conflict == 'salta' else 0,
        'n_protetti': len(protetti),
        'n_scontrini': len(risultato.scontrini),
        'n_fatture': len(risultato.fatture),
        'n_esclusi': len(risultato.esclusi),
        'strutture': sorted(risultato.strutture_trovate),
        'periodo': {
            'da': risultato.data_da.isoformat() if risultato.data_da else None,
            'a': risultato.data_a.isoformat() if risultato.data_a else None,
        },
        'warnings': risultato.warnings,
        'non_salvabili': len(risultato.righe_non_salvabili),
    }


# ── GET /corrispettivi/import/storico ─────────────────────────────────────────

@router.get("/import/storico")
def storico_import(
    is_test: bool = Query(False),
    db: Session = Depends(get_db),
    _=Depends(richiedi_utente_attivo),
):
    imports = (
        db.query(CorrispettiviImport)
        .filter(CorrispettiviImport.is_test == is_test)
        .order_by(CorrispettiviImport.created_at.desc())
        .all()
    )
    return [_fmt_import(i) for i in imports]


# ── DELETE /corrispettivi/import/{id} ─────────────────────────────────────────

@router.delete("/import/{import_id}")
def elimina_import(
    import_id: int,
    conferma: bool = Query(False),
    db: Session = Depends(get_db),
    _=Depends(richiedi_admin),
):
    """
    Elimina la sessione di import.
    Cancella i documenti collegati che non sono stati modificati manualmente.
    I documenti con modificato_manualmente=True vengono scollegati (import_id=NULL).
    """
    if not conferma:
        raise HTTPException(status_code=400,
                             detail="Aggiungere ?conferma=true per confermare l'eliminazione")

    imp = db.query(CorrispettiviImport).filter(CorrispettiviImport.id == import_id).first()
    if not imp:
        raise HTTPException(status_code=404, detail="Import non trovato")

    # Scollega i documenti modificati manualmente (audit trail da preservare)
    db.execute(
        update(CorrispettiviDocumento)
        .where(
            CorrispettiviDocumento.import_id == import_id,
            CorrispettiviDocumento.modificato_manualmente == True,
        )
        .values(import_id=None)
    )

    # Cancella i documenti non modificati manualmente
    db.query(CorrispettiviDocumento).filter(
        CorrispettiviDocumento.import_id == import_id,
        CorrispettiviDocumento.modificato_manualmente == False,
    ).delete(synchronize_session=False)

    db.delete(imp)
    db.commit()
    return {'eliminato': import_id}


# ── GET /corrispettivi/documenti ──────────────────────────────────────────────

def _lista_documenti(
    tipo: Optional[str],
    data_da: Optional[date],
    data_a: Optional[date],
    struttura_code: Optional[str],
    categoria: Optional[str],
    annullato: Optional[bool],
    is_test: bool,
    page: int,
    per_page: int,
    db: Session,
    numero: Optional[str] = None,
    camera: Optional[str] = None,
):
    q = db.query(CorrispettiviDocumento).filter(CorrispettiviDocumento.is_test == is_test)
    if tipo:
        q = q.filter(CorrispettiviDocumento.tipo == tipo)
    if data_da:
        q = q.filter(CorrispettiviDocumento.data_documento >= data_da)
    if data_a:
        q = q.filter(CorrispettiviDocumento.data_documento <= data_a)
    if struttura_code:
        q = q.filter(CorrispettiviDocumento.struttura_code == struttura_code.upper())
    if categoria:
        q = q.filter(CorrispettiviDocumento.categoria == categoria)
    if annullato is not None:
        q = q.filter(CorrispettiviDocumento.annullato == annullato)
    if numero:
        q = q.filter(CorrispettiviDocumento.numero.ilike(f'%{numero}%'))
    if camera:
        q = q.filter(CorrispettiviDocumento.camera.ilike(f'%{camera}%'))

    totale = q.count()
    totale_importo = float(
        q.with_entities(func.coalesce(func.sum(CorrispettiviDocumento.totale_lordo), 0)).scalar() or 0
    )
    docs = (
        q.order_by(CorrispettiviDocumento.data_documento.desc(),
                   CorrispettiviDocumento.numero.desc())
        .offset((page - 1) * per_page).limit(per_page).all()
    )
    return {'totale': totale, 'totale_importo': round(totale_importo, 2),
            'page': page, 'per_page': per_page,
            'documenti': [_fmt_documento(d) for d in docs]}


@router.get("/documenti")
def lista_documenti(
    tipo: Optional[str] = Query(None, description="'scontrino'|'fattura'|'escluso'|None=tutti"),
    data_da: Optional[date] = Query(None),
    data_a: Optional[date] = Query(None),
    struttura_code: Optional[str] = Query(None),
    categoria: Optional[str] = Query(None),
    annullato: Optional[bool] = Query(None),
    numero: Optional[str] = Query(None),
    camera: Optional[str] = Query(None),
    is_test: bool = Query(False),
    page: int = Query(1, ge=1),
    per_page: int = Query(50, ge=1, le=200),
    db: Session = Depends(get_db),
    _=Depends(richiedi_utente_attivo),
):
    return _lista_documenti(tipo, data_da, data_a, struttura_code, categoria,
                             annullato, is_test, page, per_page, db, numero, camera)


@router.get("/scontrini")
def lista_scontrini(
    data_da: Optional[date] = Query(None),
    data_a: Optional[date] = Query(None),
    struttura_code: Optional[str] = Query(None),
    categoria: Optional[str] = Query(None),
    annullato: Optional[bool] = Query(None),
    numero: Optional[str] = Query(None),
    camera: Optional[str] = Query(None),
    is_test: bool = Query(False),
    page: int = Query(1, ge=1),
    per_page: int = Query(50, ge=1, le=200),
    db: Session = Depends(get_db),
    _=Depends(richiedi_utente_attivo),
):
    return _lista_documenti('scontrino', data_da, data_a, struttura_code, categoria,
                             annullato, is_test, page, per_page, db, numero, camera)


@router.get("/fatture")
def lista_fatture(
    data_da: Optional[date] = Query(None),
    data_a: Optional[date] = Query(None),
    struttura_code: Optional[str] = Query(None),
    categoria: Optional[str] = Query(None),
    annullato: Optional[bool] = Query(None),
    numero: Optional[str] = Query(None),
    camera: Optional[str] = Query(None),
    is_test: bool = Query(False),
    page: int = Query(1, ge=1),
    per_page: int = Query(50, ge=1, le=200),
    db: Session = Depends(get_db),
    _=Depends(richiedi_utente_attivo),
):
    return _lista_documenti('fattura', data_da, data_a, struttura_code, categoria,
                             annullato, is_test, page, per_page, db, numero, camera)


# ── PUT /corrispettivi/documenti/{id} ─────────────────────────────────────────

def _modifica_documento(doc_id: int, body: dict, db: Session, utente) -> dict:
    doc = db.query(CorrispettiviDocumento).filter(CorrispettiviDocumento.id == doc_id).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Documento non trovato")

    # Salva valori originali al primo edit manuale
    if not doc.modificato_manualmente:
        doc.totale_lordo_originale = doc.totale_lordo
        doc.imponibile_originale = doc.imponibile
        doc.iva_originale = doc.iva
        doc.categoria_originale = doc.categoria

    campi_modificabili = {
        'totale_lordo', 'incassato', 'deposito', 'sospeso', 'abbuono',
        'imponibile', 'iva', 'categoria', 'annullato', 'note', 'ospiti', 'modifica_note',
        'tipo_pagamento', 'categoria_pagamento',
    }
    for campo, valore in body.items():
        if campo in campi_modificabili:
            setattr(doc, campo, valore)

    # Ricalcola aliquota e categoria se imponibile/iva modificati
    if 'imponibile' in body or 'iva' in body:
        imponibile = float(doc.imponibile or 0)
        iva = float(doc.iva or 0)
        if imponibile:
            doc.aliquota_pct = round(iva / imponibile * 100, 2)
        if 'categoria' not in body:
            doc.categoria = _determina_categoria(float(doc.aliquota_pct), imponibile)

    doc.modificato_manualmente = True
    doc.modificato_da = utente.id
    doc.modificato_at = datetime.utcnow()

    db.commit()
    return _fmt_documento(doc)


@router.put("/documenti/{doc_id}")
def modifica_documento(
    doc_id: int,
    body: dict,
    db: Session = Depends(get_db),
    utente=Depends(richiedi_admin),
):
    """Correzione manuale unificata. Salva i valori originali al primo edit."""
    return _modifica_documento(doc_id, body, db, utente)


@router.put("/scontrini/{doc_id}")
def modifica_scontrino(
    doc_id: int,
    body: dict,
    db: Session = Depends(get_db),
    utente=Depends(richiedi_admin),
):
    return _modifica_documento(doc_id, body, db, utente)


@router.put("/fatture/{doc_id}")
def modifica_fattura(
    doc_id: int,
    body: dict,
    db: Session = Depends(get_db),
    utente=Depends(richiedi_admin),
):
    return _modifica_documento(doc_id, body, db, utente)


# ── POST /corrispettivi/manuali ───────────────────────────────────────────────

@router.post("/manuali")
def crea_manuale(
    body: dict,
    db: Session = Depends(get_db),
    utente=Depends(richiedi_admin),
):
    """
    Inserisce o aggiorna un corrispettivo manuale per MMS o BON.
    Body: {data_giorno, struttura_code, arrangiamenti_lordo, note?, is_test?}
    """
    struttura = str(body.get('struttura_code', '')).upper()
    if struttura not in STRUTTURE_MANUALI:
        raise HTTPException(status_code=400,
                             detail=f"struttura_code deve essere uno di: {STRUTTURE_MANUALI}")

    data_g_raw = body.get('data_giorno')
    if not data_g_raw:
        raise HTTPException(status_code=400, detail="data_giorno obbligatorio")
    try:
        data_g = date.fromisoformat(str(data_g_raw))
    except ValueError:
        raise HTTPException(status_code=400, detail="data_giorno non valida (formato YYYY-MM-DD)")

    lordo = float(body.get('arrangiamenti_lordo', 0))
    is_test = bool(body.get('is_test', False))

    esistente = db.query(CorrispettiviManuale).filter(
        CorrispettiviManuale.data_giorno == data_g,
        CorrispettiviManuale.struttura_code == struttura,
    ).first()

    if esistente:
        esistente.arrangiamenti_lordo = _d(lordo)
        esistente.note = body.get('note', esistente.note)
        esistente.updated_by = utente.id
        m = esistente
    else:
        m = CorrispettiviManuale(
            data_giorno=data_g,
            struttura_code=struttura,
            arrangiamenti_lordo=_d(lordo),
            note=body.get('note'),
            is_test=is_test,
            created_by=utente.id,
            updated_by=utente.id,
        )
        db.add(m)

    db.commit()
    return _fmt_manuale(m)


# ── PUT /corrispettivi/manuali/{id} ───────────────────────────────────────────

@router.put("/manuali/{manuale_id}")
def modifica_manuale(
    manuale_id: int,
    body: dict,
    db: Session = Depends(get_db),
    utente=Depends(richiedi_admin),
):
    m = db.query(CorrispettiviManuale).filter(CorrispettiviManuale.id == manuale_id).first()
    if not m:
        raise HTTPException(status_code=404, detail="Record non trovato")

    if 'arrangiamenti_lordo' in body:
        m.arrangiamenti_lordo = _d(body['arrangiamenti_lordo'])
    if 'note' in body:
        m.note = body['note']
    m.updated_by = utente.id

    db.commit()
    return _fmt_manuale(m)


# ── GET /corrispettivi/manuali ────────────────────────────────────────────────

@router.get("/manuali")
def lista_manuali(
    data_da: Optional[date] = Query(None),
    data_a: Optional[date] = Query(None),
    struttura_code: Optional[str] = Query(None),
    is_test: bool = Query(False),
    db: Session = Depends(get_db),
    _=Depends(richiedi_utente_attivo),
):
    q = db.query(CorrispettiviManuale).filter(CorrispettiviManuale.is_test == is_test)
    if data_da:
        q = q.filter(CorrispettiviManuale.data_giorno >= data_da)
    if data_a:
        q = q.filter(CorrispettiviManuale.data_giorno <= data_a)
    if struttura_code:
        q = q.filter(CorrispettiviManuale.struttura_code == struttura_code.upper())
    manuali = q.order_by(CorrispettiviManuale.data_giorno.desc()).all()
    return [_fmt_manuale(m) for m in manuali]


# ── GET /corrispettivi/report/giornaliero ─────────────────────────────────────

@router.get("/report/giornaliero")
def report_giornaliero(
    data_da: Optional[date] = Query(None),
    data_a: Optional[date] = Query(None),
    struttura_code: Optional[str] = Query(None),
    tipo: str = Query('tutti', description="'scontrini'|'fatture'|'tutti'"),
    lordo: bool = Query(True),
    is_test: bool = Query(False),
    db: Session = Depends(get_db),
    _=Depends(richiedi_utente_attivo),
):
    """Aggregato giornaliero per struttura con dettaglio per categoria."""

    # Aliquote per categoria (per conversione lordo→netto)
    ALIQ: dict = {
        'arrangiamenti': 10.0,
        'tassa_soggiorno': 0.0,
        'penali': 0.0,
        'shop': 22.0,
        'altro': 10.0,
    }

    def _v(val_lordo: float, cat: str) -> float:
        if lordo:
            return round(val_lordo, 2)
        aliq = ALIQ.get(cat, 10.0)
        return round(val_lordo / (1 + aliq / 100), 2) if aliq else round(val_lordo, 2)

    # Filtra tipi documenti da includere
    tipi_filtro: List[str] = []
    if tipo == 'scontrini':
        tipi_filtro = ['scontrino']
    elif tipo == 'fatture':
        tipi_filtro = ['fattura']
    else:
        tipi_filtro = ['scontrino', 'fattura']

    # Aggrega da corrispettivi_documenti
    q = (
        db.query(
            CorrispettiviDocumento.data_documento,
            CorrispettiviDocumento.struttura_code,
            CorrispettiviDocumento.tipo,
            CorrispettiviDocumento.categoria,
            func.sum(CorrispettiviDocumento.totale_lordo).label('totale_lordo'),
            func.sum(CorrispettiviDocumento.iva).label('sum_iva'),
            # Somma tassa_soggiorno (NULL se colonna assente nel file sorgente)
            func.sum(CorrispettiviDocumento.tassa_soggiorno).label('sum_ts'),
            func.count().label('n_tot'),
            func.count().filter(CorrispettiviDocumento.annullato == True).label('n_annullati'),
        )
        .filter(
            CorrispettiviDocumento.tipo.in_(tipi_filtro),
            CorrispettiviDocumento.is_test == is_test,
        )
    )
    if data_da:
        q = q.filter(CorrispettiviDocumento.data_documento >= data_da)
    if data_a:
        q = q.filter(CorrispettiviDocumento.data_documento <= data_a)
    if struttura_code:
        q = q.filter(CorrispettiviDocumento.struttura_code == struttura_code.upper())

    righe_doc = q.group_by(
        CorrispettiviDocumento.data_documento,
        CorrispettiviDocumento.struttura_code,
        CorrispettiviDocumento.tipo,
        CorrispettiviDocumento.categoria,
    ).all()

    # Aggrega manuali
    q_man = db.query(CorrispettiviManuale).filter(CorrispettiviManuale.is_test == is_test)
    if data_da:
        q_man = q_man.filter(CorrispettiviManuale.data_giorno >= data_da)
    if data_a:
        q_man = q_man.filter(CorrispettiviManuale.data_giorno <= data_a)
    if struttura_code and struttura_code.upper() not in STRUTTURE_MANUALI:
        manuali_rows = []
    else:
        if struttura_code:
            q_man = q_man.filter(CorrispettiviManuale.struttura_code == struttura_code.upper())
        manuali_rows = q_man.all()

    # Struttura dati: date → struttura → tipo → categoria → totale_lordo
    per_data: dict = {}

    for r in righe_doc:
        dkey = r.data_documento.isoformat()
        if dkey not in per_data:
            per_data[dkey] = {}
        sc = r.struttura_code
        if sc not in per_data[dkey]:
            per_data[dkey][sc] = {
                'scontrino': {'cat': {c: 0.0 for c in CATEGORIE}, 'n': 0, 'n_ann': 0},
                'fattura':   {'cat': {c: 0.0 for c in CATEGORIE}, 'n': 0, 'n_ann': 0},
            }
        t = r.tipo
        cat = r.categoria or 'altro'
        non_ann = int(r.n_tot or 0) - int(r.n_annullati or 0)
        lordo_tot = _to_float(r.totale_lordo)
        sum_iva = _to_float(r.sum_iva)
        # sum_ts: None = colonna tassa_soggiorno assente nel file (vecchio formato)
        #         float = valore esatto disponibile (nuovo formato esteso)
        sum_ts = None if r.sum_ts is None else float(r.sum_ts)

        # Disaggregazione mix arrangiamenti + tassa soggiorno per documenti con cat in
        # ('arrangiamenti', 'altro'): aliquota effettiva < 10% indica presenza di TS esente
        if cat in ('arrangiamenti', 'altro'):
            if sum_ts is not None:
                # Nuovo formato: valore esatto dalla colonna 'Tassa di soggiorno'
                lordo_ts = round(sum_ts, 2)
                lordo_arr = round(max(0.0, lordo_tot - lordo_ts), 2)
                per_data[dkey][sc][t]['cat']['arrangiamenti'] += lordo_arr
                if lordo_ts > 0:
                    per_data[dkey][sc][t]['cat']['tassa_soggiorno'] += lordo_ts
            elif sum_iva > 0:
                # Fallback formato base: inferenza da IVA (imp=iva*10, lordo_arr=iva*11)
                lordo_arr = round(sum_iva * 11, 2)
                lordo_ts = round(max(0.0, lordo_tot - lordo_arr), 2)
                per_data[dkey][sc][t]['cat']['arrangiamenti'] += lordo_arr
                if lordo_ts > 0:
                    per_data[dkey][sc][t]['cat']['tassa_soggiorno'] += lordo_ts
            else:
                per_data[dkey][sc][t]['cat'][cat] += lordo_tot
        else:
            per_data[dkey][sc][t]['cat'][cat] += lordo_tot

        per_data[dkey][sc][t]['n'] += non_ann
        per_data[dkey][sc][t]['n_ann'] += int(r.n_annullati or 0)

    # Integra manuali
    per_data_manuali: dict = {}
    for m in manuali_rows:
        dkey = m.data_giorno.isoformat()
        if dkey not in per_data_manuali:
            per_data_manuali[dkey] = {}
        per_data_manuali[dkey][m.struttura_code] = _to_float(m.arrangiamenti_lordo)

    # Raccoglie tutte le date presenti
    tutte_date = sorted(set(list(per_data.keys()) + list(per_data_manuali.keys())))

    risultato = []
    for dkey in tutte_date:
        strutture_giorno = []
        totale_giorno = 0.0

        strutture_in_giorno: Set[str] = set()
        if dkey in per_data:
            strutture_in_giorno.update(per_data[dkey].keys())
        if dkey in per_data_manuali:
            strutture_in_giorno.update(per_data_manuali[dkey].keys())

        for sc in STRUTTURE_ORDINE:
            if sc not in strutture_in_giorno:
                continue

            dati_sc = per_data.get(dkey, {}).get(sc, {
                'scontrino': {'cat': {c: 0.0 for c in CATEGORIE}, 'n': 0, 'n_ann': 0},
                'fattura':   {'cat': {c: 0.0 for c in CATEGORIE}, 'n': 0, 'n_ann': 0},
            })

            def _cat_dict(t: str) -> dict:
                cats = dati_sc[t]['cat']
                return {
                    cat: _v(cats.get(cat, 0.0), cat)
                    for cat in CATEGORIE
                }

            sc_cats = _cat_dict('scontrino')
            f_cats  = _cat_dict('fattura')

            sc_tot = round(sum(sc_cats.values()), 2)
            f_tot  = round(sum(f_cats.values()), 2)

            man_lordo = _to_float(per_data_manuali.get(dkey, {}).get(sc, 0))
            man = _v(man_lordo, 'arrangiamenti') if sc in STRUTTURE_MANUALI else 0.0

            totale_struttura = sc_tot + f_tot + man
            totale_giorno += totale_struttura

            strutture_giorno.append({
                'struttura_code': sc,
                'struttura_nome': NOME_STRUTTURA.get(sc, sc),
                'scontrini': {**sc_cats, 'totale': sc_tot},
                'fatture':   {**f_cats,  'totale': f_tot},
                'manuale': man,
                'totale': round(totale_struttura, 2),
                'n_scontrini': dati_sc['scontrino']['n'],
                'n_fatture':   dati_sc['fattura']['n'],
                'n_annullati': dati_sc['scontrino']['n_ann'] + dati_sc['fattura']['n_ann'],
            })

        risultato.append({
            'data': dkey,
            'strutture': strutture_giorno,
            'totale_giorno': round(totale_giorno, 2),
        })

    return risultato


# ── GET /corrispettivi/report/mensile ─────────────────────────────────────────

@router.get("/report/mensile")
def report_mensile(
    mese: int = Query(..., ge=1, le=12),
    anno: int = Query(..., ge=2020, le=2030),
    struttura_code: Optional[str] = Query(None),
    tipo: str = Query('tutti'),
    lordo: bool = Query(True),
    is_test: bool = Query(False),
    db: Session = Depends(get_db),
    _=Depends(richiedi_utente_attivo),
):
    import calendar
    primo = date(anno, mese, 1)
    ultimo = date(anno, mese, calendar.monthrange(anno, mese)[1])

    risposta_gg = report_giornaliero(
        data_da=primo, data_a=ultimo,
        struttura_code=struttura_code,
        tipo=tipo, lordo=lordo, is_test=is_test,
        db=db, _=None,
    )

    aggregati: dict = {}
    for giorno in risposta_gg:
        for s in giorno['strutture']:
            sc = s['struttura_code']
            if sc not in aggregati:
                aggregati[sc] = {
                    'struttura_code': sc,
                    'struttura_nome': s['struttura_nome'],
                    'scontrini': {c: 0.0 for c in CATEGORIE + ['totale']},
                    'fatture':   {c: 0.0 for c in CATEGORIE + ['totale']},
                    'manuale': 0.0, 'totale': 0.0,
                    'n_scontrini': 0, 'n_fatture': 0, 'n_annullati': 0,
                }
            agg = aggregati[sc]
            for cat in CATEGORIE + ['totale']:
                agg['scontrini'][cat] += s['scontrini'].get(cat, 0)
                agg['fatture'][cat]   += s['fatture'].get(cat, 0)
            agg['manuale']      += s['manuale']
            agg['totale']       += s['totale']
            agg['n_scontrini']  += s['n_scontrini']
            agg['n_fatture']    += s['n_fatture']
            agg['n_annullati']  += s['n_annullati']

    strutture_out = [aggregati[sc] for sc in STRUTTURE_ORDINE if sc in aggregati]
    for sc in sorted(aggregati):
        if sc not in STRUTTURE_ORDINE and sc not in [s['struttura_code'] for s in strutture_out]:
            strutture_out.append(aggregati[sc])

    totale_generale = sum(s['totale'] for s in strutture_out)
    return {
        'mese': mese,
        'anno': anno,
        'strutture': strutture_out,
        'totale_generale': round(totale_generale, 2),
    }


# ── GET /corrispettivi/check ──────────────────────────────────────────────────

@router.get("/check")
def check_totali(
    data_da: Optional[date] = Query(None),
    data_a: Optional[date] = Query(None),
    lordo: bool = Query(True),
    is_test: bool = Query(False),
    db: Session = Depends(get_db),
    _=Depends(richiedi_utente_attivo),
):
    """Totali aggregati per struttura separando hotel e ristoranti."""
    risposta_gg = report_giornaliero(
        data_da=data_da, data_a=data_a,
        struttura_code=None, tipo='tutti',
        lordo=lordo, is_test=is_test,
        db=db, _=None,
    )

    per_struttura: dict = {s: 0.0 for s in STRUTTURE_ORDINE}
    for giorno in risposta_gg:
        for s in giorno['strutture']:
            sc = s['struttura_code']
            if sc in per_struttura:
                per_struttura[sc] += s['totale']

    totale_hotel = sum(per_struttura.get(s, 0) for s in STRUTTURE_HOTEL)
    totale_ristoranti = sum(per_struttura.get(s, 0) for s in STRUTTURE_MANUALI)
    totale_generale = totale_hotel + totale_ristoranti

    return {
        **{sc: round(per_struttura.get(sc, 0), 2) for sc in STRUTTURE_ORDINE},
        'totale_hotel': round(totale_hotel, 2),
        'totale_ristoranti': round(totale_ristoranti, 2),
        'totale_generale': round(totale_generale, 2),
        'label_hotel': ' + '.join(STRUTTURE_HOTEL),
        'label_ristoranti': ' + '.join(STRUTTURE_MANUALI),
    }


# ── GET /corrispettivi/report/fatturati ───────────────────────────────────────

@router.get("/report/fatturati")
def report_fatturati(
    anno: int = Query(..., ge=2020, le=2030),
    lordo: bool = Query(True),
    is_test: bool = Query(False),
    db: Session = Depends(get_db),
    _=Depends(richiedi_utente_attivo),
):
    """Riepilogo fatturati per mese e struttura nell'anno.

    Aggrega scontrini e fatture da corrispettivi_documenti
    (annullato=false) più corrispettivi_manuali per MMS/BON.
    Tassa soggiorno esclusa dal totale (transito verso il Comune).
    Se lordo=false applica divisori per categoria.
    """
    from app.utils.locale_it import MESI_IT

    ALIQ: dict = {
        'arrangiamenti': 10.0,
        'tassa_soggiorno': 0.0,
        'penali': 0.0,
        'shop': 22.0,
        'altro': 10.0,
    }

    def _v(val_lordo: float, cat: str) -> float:
        if lordo:
            return round(val_lordo, 2)
        aliq = ALIQ.get(cat, 10.0)
        return round(val_lordo / (1 + aliq / 100), 2) if aliq else round(val_lordo, 2)

    # ── Query documenti aggregata per (mese, struttura, categoria) ────────────
    mese_col = func.extract('month', CorrispettiviDocumento.data_documento).label('mese')
    righe_doc = (
        db.query(
            mese_col,
            CorrispettiviDocumento.struttura_code,
            CorrispettiviDocumento.categoria,
            func.sum(CorrispettiviDocumento.totale_lordo).label('totale_lordo'),
            func.sum(CorrispettiviDocumento.iva).label('sum_iva'),
            func.sum(CorrispettiviDocumento.tassa_soggiorno).label('sum_ts'),
        )
        .filter(
            func.extract('year', CorrispettiviDocumento.data_documento) == anno,
            CorrispettiviDocumento.tipo.in_(['scontrino', 'fattura']),
            CorrispettiviDocumento.annullato == False,
            CorrispettiviDocumento.is_test == is_test,
        )
        .group_by(mese_col, CorrispettiviDocumento.struttura_code, CorrispettiviDocumento.categoria)
        .all()
    )

    # ── Query manuali aggregata per (mese, struttura) ─────────────────────────
    mese_man_col = func.extract('month', CorrispettiviManuale.data_giorno).label('mese')
    righe_man = (
        db.query(
            mese_man_col,
            CorrispettiviManuale.struttura_code,
            func.sum(CorrispettiviManuale.arrangiamenti_lordo).label('arrangiamenti_lordo'),
        )
        .filter(
            func.extract('year', CorrispettiviManuale.data_giorno) == anno,
            CorrispettiviManuale.is_test == is_test,
        )
        .group_by(mese_man_col, CorrispettiviManuale.struttura_code)
        .all()
    )

    # ── Accumulo per (mese, struttura) ────────────────────────────────────────
    # struttura → {arrangiamenti, tassa_soggiorno, penali, shop, altro}
    def _empty_cats() -> dict:
        return {c: 0.0 for c in CATEGORIE}

    per_mese: dict = {}  # mese(int) → struttura → cats

    for r in righe_doc:
        m = int(r.mese)
        sc = r.struttura_code
        cat = r.categoria or 'altro'
        lordo_tot = _to_float(r.totale_lordo)
        sum_iva = _to_float(r.sum_iva)
        sum_ts = None if r.sum_ts is None else float(r.sum_ts)

        if m not in per_mese:
            per_mese[m] = {}
        if sc not in per_mese[m]:
            per_mese[m][sc] = _empty_cats()

        # Stessa disaggregazione usata nel report giornaliero
        if cat in ('arrangiamenti', 'altro'):
            if sum_ts is not None:
                lordo_ts = round(sum_ts, 2)
                lordo_arr = round(max(0.0, lordo_tot - lordo_ts), 2)
                per_mese[m][sc]['arrangiamenti'] += lordo_arr
                if lordo_ts > 0:
                    per_mese[m][sc]['tassa_soggiorno'] += lordo_ts
            elif sum_iva > 0:
                lordo_arr = round(sum_iva * 11, 2)
                lordo_ts = round(max(0.0, lordo_tot - lordo_arr), 2)
                per_mese[m][sc]['arrangiamenti'] += lordo_arr
                if lordo_ts > 0:
                    per_mese[m][sc]['tassa_soggiorno'] += lordo_ts
            else:
                per_mese[m][sc][cat] += lordo_tot
        else:
            per_mese[m][sc][cat] += lordo_tot

    for r in righe_man:
        m = int(r.mese)
        sc = r.struttura_code
        arr = _to_float(r.arrangiamenti_lordo)
        if m not in per_mese:
            per_mese[m] = {}
        if sc not in per_mese[m]:
            per_mese[m][sc] = _empty_cats()
        per_mese[m][sc]['arrangiamenti'] += arr

    if not per_mese:
        return {
            'anno': anno,
            'strutture': [],
            'mesi': [],
            'totale_anno': {
                'per_struttura': {},
                'totale_hotel': 0.0,
                'totale_ristoranti': 0.0,
                'totale_generale': 0.0,
            },
        }

    # Strutture presenti nell'anno, ordinate per STRUTTURE_ORDINE
    strutture_presenti = [
        s for s in STRUTTURE_ORDINE
        if any(s in per_mese[m] for m in per_mese)
    ]

    def _struttura_output(cats_lordo: dict) -> dict:
        """Converte i valori lordi in output (lordo o netto), calcola totale senza TS."""
        out = {c: _v(cats_lordo.get(c, 0.0), c) for c in CATEGORIE}
        # totale esclude tassa_soggiorno e altro
        out['totale'] = round(
            out['arrangiamenti'] + out['penali'] + out['shop'] + out['altro'], 2
        )
        return out

    # ── Totale anno per struttura ─────────────────────────────────────────────
    totale_anno_cats: dict = {s: _empty_cats() for s in strutture_presenti}
    for m, per_sc in per_mese.items():
        for sc, cats in per_sc.items():
            if sc in totale_anno_cats:
                for c in CATEGORIE:
                    totale_anno_cats[sc][c] += cats.get(c, 0.0)

    totale_anno_per_struttura = {
        sc: _struttura_output(totale_anno_cats[sc])
        for sc in strutture_presenti
    }
    tot_anno_hotel = round(sum(
        totale_anno_per_struttura[s]['totale']
        for s in strutture_presenti if s in STRUTTURE_HOTEL
    ), 2)
    tot_anno_rist = round(sum(
        totale_anno_per_struttura[s]['totale']
        for s in strutture_presenti if s in STRUTTURE_MANUALI
    ), 2)

    # ── Costruzione risposta mesi ─────────────────────────────────────────────
    mesi_output = []
    for m in sorted(per_mese.keys()):
        per_sc = per_mese[m]
        per_struttura_out = {}
        for sc in strutture_presenti:
            cats = per_sc.get(sc, _empty_cats())
            per_struttura_out[sc] = _struttura_output(cats)

        tot_hotel = round(sum(
            per_struttura_out[s]['totale']
            for s in strutture_presenti if s in STRUTTURE_HOTEL
        ), 2)
        tot_rist = round(sum(
            per_struttura_out[s]['totale']
            for s in strutture_presenti if s in STRUTTURE_MANUALI
        ), 2)

        mesi_output.append({
            'mese': m,
            'nome_mese': MESI_IT[m - 1],
            'per_struttura': per_struttura_out,
            'totale_hotel': tot_hotel,
            'totale_ristoranti': tot_rist,
            'totale_generale': round(tot_hotel + tot_rist, 2),
        })

    return {
        'anno': anno,
        'strutture': strutture_presenti,
        'mesi': mesi_output,
        'totale_anno': {
            'per_struttura': totale_anno_per_struttura,
            'totale_hotel': tot_anno_hotel,
            'totale_ristoranti': tot_anno_rist,
            'totale_generale': round(tot_anno_hotel + tot_anno_rist, 2),
        },
    }


# ── GET /corrispettivi/report/pagamenti ──────────────────────────────────────

@router.get("/report/pagamenti")
def report_pagamenti(
    anno: int = Query(..., ge=2020, le=2030),
    is_test: bool = Query(False),
    db: Session = Depends(get_db),
    _=Depends(richiedi_utente_attivo),
):
    """Riepilogo fatturati per tipo di pagamento e mese nell'anno.

    Normalizza il testo grezzo della colonna Pagamenti (es. 'Contante 8,00 € /')
    estraendo solo il nome del tipo.  Restituisce sempre valori lordi.
    """
    from app.utils.locale_it import MESI_IT

    TIPI_NOTI = [
        'Bonifico/Vaglia', 'XPAY-Nexi', 'Carta Credito',
        'Bancomat', 'Bonifico', 'Contante', 'Satispay', 'Assegno', 'xpay',
    ]

    def _normalizza(raw: str) -> str:
        if not raw or not raw.strip():
            return 'Non specificato'
        raw_l = raw.strip().lower()
        for tipo in TIPI_NOTI:
            if raw_l.startswith(tipo.lower()):
                return tipo
        return raw.strip()

    mese_col = func.extract('month', CorrispettiviDocumento.data_documento).label('mese')
    # Ogni documento si divide in tre componenti:
    #   pagato (tipo_pagamento) = totale_lordo - sospeso - deposito - tassa_soggiorno
    #   caparra (deposito)      = deposito
    #   sospeso                 = sospeso
    # Totale = pagato + caparra + sospeso = totale_lordo - TS  →  pareggia con report_fatturati.
    from sqlalchemy import func as sqlfunc
    pagato_col  = func.sum(
        CorrispettiviDocumento.totale_lordo
        - CorrispettiviDocumento.sospeso
        - CorrispettiviDocumento.deposito
        - sqlfunc.coalesce(CorrispettiviDocumento.tassa_soggiorno, 0)
    ).label('totale_pagato')
    sospeso_col = func.sum(CorrispettiviDocumento.sospeso).label('totale_sospeso')
    deposito_col = func.sum(CorrispettiviDocumento.deposito).label('totale_deposito')

    righe = (
        db.query(mese_col, CorrispettiviDocumento.tipo_pagamento,
                 pagato_col, sospeso_col, deposito_col)
        .filter(
            func.extract('year', CorrispettiviDocumento.data_documento) == anno,
            CorrispettiviDocumento.tipo.in_(['scontrino', 'fattura']),
            CorrispettiviDocumento.annullato == False,
            CorrispettiviDocumento.is_test == is_test,
        )
        .group_by(mese_col, CorrispettiviDocumento.tipo_pagamento)
        .all()
    )

    # Aggiungi i corrispettivi manuali MMS/BON per mese
    mese_man_col = func.extract('month', CorrispettiviManuale.data_giorno).label('mese')
    righe_man = (
        db.query(mese_man_col, func.sum(CorrispettiviManuale.arrangiamenti_lordo).label('totale'))
        .filter(
            func.extract('year', CorrispettiviManuale.data_giorno) == anno,
            CorrispettiviManuale.is_test == is_test,
        )
        .group_by(mese_man_col)
        .all()
    )

    if not righe and not righe_man:
        return {'anno': anno, 'mesi': [], 'tipi': [], 'per_tipo': {},
                'totale_mese': {}, 'totale_anno': {}}

    per_tipo: dict = {}
    mesi_set: set = set()

    def _acc(key: str, mese: int, val: float):
        if val == 0:
            return
        if key not in per_tipo:
            per_tipo[key] = {}
        per_tipo[key][mese] = round(per_tipo[key].get(mese, 0.0) + val, 2)

    for r in righe:
        m = int(r.mese)
        mesi_set.add(m)

        val_pag  = round(float(r.totale_pagato  or 0), 2)
        val_sosp = round(float(r.totale_sospeso or 0), 2)
        val_dep  = round(float(r.totale_deposito or 0), 2)

        if val_pag != 0:
            _acc(_normalizza(str(r.tipo_pagamento or '')), m, val_pag)
        _acc('Caparra', m, val_dep)
        _acc('Sospeso', m, val_sosp)

    for r in righe_man:
        m = int(r.mese)
        mesi_set.add(m)
        _acc('MMS / BON (manuale)', m, round(float(r.totale or 0), 2))

    mesi_ordinati = sorted(mesi_set)

    # Ordine: tipi noti → altri → Non specificato → MMS/BON → Caparra → Sospeso in fondo
    SPECIALI = ('Non specificato', 'MMS / BON (manuale)', 'Caparra', 'Sospeso')
    tipi_noti_presenti = [t for t in TIPI_NOTI if t in per_tipo]
    tipi_altri = sorted(t for t in per_tipo if t not in TIPI_NOTI and t not in SPECIALI)
    tipi_ns    = ['Non specificato']    if 'Non specificato'    in per_tipo else []
    tipi_man   = ['MMS / BON (manuale)'] if 'MMS / BON (manuale)' in per_tipo else []
    tipi_cap   = ['Caparra']            if 'Caparra'            in per_tipo else []
    tipi_sosp  = ['Sospeso']            if 'Sospeso'            in per_tipo else []
    tipi_ordinati = tipi_noti_presenti + tipi_altri + tipi_ns + tipi_man + tipi_cap + tipi_sosp

    totale_mese = {
        m: round(sum(per_tipo[t].get(m, 0.0) for t in tipi_ordinati), 2)
        for m in mesi_ordinati
    }
    totale_anno = {t: round(sum(per_tipo[t].values()), 2) for t in tipi_ordinati}

    return {
        'anno': anno,
        'mesi': [{'mese': m, 'nome_mese': MESI_IT[m - 1]} for m in mesi_ordinati],
        'tipi': tipi_ordinati,
        'per_tipo': {
            t: {str(m): per_tipo[t].get(m, 0.0) for m in mesi_ordinati}
            for t in tipi_ordinati
        },
        'totale_mese': {str(m): totale_mese[m] for m in mesi_ordinati},
        'totale_anno': totale_anno,
    }


# ── Admin: test data ──────────────────────────────────────────────────────────

@router.get("/admin/test-stats")
def test_stats(
    db: Session = Depends(get_db),
    _=Depends(richiedi_admin),
):
    n_doc = db.query(func.count(CorrispettiviDocumento.id)).filter(
        CorrispettiviDocumento.is_test == True).scalar() or 0
    n_imp = db.query(func.count(CorrispettiviImport.id)).filter(
        CorrispettiviImport.is_test == True).scalar() or 0
    n_man = db.query(func.count(CorrispettiviManuale.id)).filter(
        CorrispettiviManuale.is_test == True).scalar() or 0
    return {
        'imports': n_imp,
        'documenti': n_doc,
        'manuali': n_man,
        'totale': n_imp + n_doc + n_man,
    }


@router.delete("/admin/test-data")
def elimina_test_data(
    conferma: bool = Query(False),
    db: Session = Depends(get_db),
    _=Depends(richiedi_admin),
):
    if not conferma:
        raise HTTPException(status_code=400,
                             detail="Aggiungere ?conferma=true per confermare")

    db.query(CorrispettiviDocumento).filter(
        CorrispettiviDocumento.is_test == True).delete(synchronize_session=False)
    db.query(CorrispettiviManuale).filter(
        CorrispettiviManuale.is_test == True).delete(synchronize_session=False)
    db.query(CorrispettiviImport).filter(
        CorrispettiviImport.is_test == True).delete(synchronize_session=False)

    db.commit()
    return {'eliminati': True}


# ── GET /corrispettivi/export/fatturati ───────────────────────────────────────

@router.get("/export/fatturati")
def export_fatturati(
    anno: int = Query(..., ge=2020, le=2030),
    lordo: bool = Query(True),
    is_test: bool = Query(False),
    db: Session = Depends(get_db),
    _=Depends(richiedi_utente_attivo),
):
    """Esporta il riepilogo fatturati in Excel (3 fogli: Corrispettivi, Tassa Soggiorno, Controllo)."""
    import io
    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from app.utils.locale_it import MESI_IT

    # Riusa la logica di report_fatturati
    dati = report_fatturati(anno=anno, lordo=lordo, is_test=is_test, db=db, _=None)
    strutture = dati['strutture']
    mesi = dati['mesi']
    tot_anno = dati['totale_anno']

    strutture_hotel = [s for s in strutture if s in STRUTTURE_HOTEL]

    HDR_FILL = PatternFill('solid', fgColor='1E3A5F')
    HDR_FONT = Font(bold=True, color='FFFFFF', size=9)
    TOT_FILL = PatternFill('solid', fgColor='1E3A5F')
    TOT_FONT = Font(bold=True, color='FFFFFF', size=9)
    ALT_FILL = PatternFill('solid', fgColor='F8FAFC')
    TEAL_FILL = PatternFill('solid', fgColor='0F766E')
    NUM_FMT = '#,##0.00 "€"'

    wb = Workbook()

    def _hdr(ws, row, cols):
        for c, val in enumerate(cols, 1):
            cell = ws.cell(row=row, column=c, value=val)
            cell.font = HDR_FONT
            cell.fill = HDR_FILL
            cell.alignment = Alignment(horizontal='center')

    def _tot_row(ws, row, vals):
        for c, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=c, value=val)
            cell.font = TOT_FONT
            cell.fill = TOT_FILL
            cell.alignment = Alignment(horizontal='right' if c > 1 else 'left')
            if isinstance(val, float):
                cell.number_format = NUM_FMT

    def _data_row(ws, row, vals, alt=False):
        fill = ALT_FILL if alt else None
        for c, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=c, value=val)
            cell.alignment = Alignment(horizontal='right' if c > 1 else 'left')
            if fill:
                cell.fill = fill
            if isinstance(val, float):
                cell.number_format = NUM_FMT

    def _autowidth(ws):
        for col in ws.columns:
            max_len = max((len(str(cell.value or '')) for cell in col), default=8)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 22)

    # ── Foglio 1: Totale Corrispettivi ────────────────────────────────────────
    ws1 = wb.active
    ws1.title = 'Corrispettivi'
    hdrs = ['Mese'] + strutture + ['Tot. Hotel', 'Tot. Rist.', 'TOTALE']
    _hdr(ws1, 1, hdrs)
    for i, m in enumerate(mesi):
        row = i + 2
        vals = [m['nome_mese']]
        for s in strutture:
            vals.append(m['per_struttura'].get(s, {}).get('totale') or 0.0)
        vals += [m.get('totale_hotel') or 0.0,
                 m.get('totale_ristoranti') or 0.0,
                 m.get('totale_generale') or 0.0]
        _data_row(ws1, row, vals, alt=i % 2 == 1)
    tot_vals = ['TOTALE ANNO']
    for s in strutture:
        tot_vals.append(tot_anno['per_struttura'].get(s, {}).get('totale') or 0.0)
    tot_vals += [tot_anno.get('totale_hotel') or 0.0,
                 tot_anno.get('totale_ristoranti') or 0.0,
                 tot_anno.get('totale_generale') or 0.0]
    _tot_row(ws1, len(mesi) + 2, tot_vals)
    _autowidth(ws1)

    # ── Foglio 2: Tassa di Soggiorno ──────────────────────────────────────────
    ws2 = wb.create_sheet('Tassa di Soggiorno')
    hdrs2 = ['Mese'] + strutture_hotel + ['Tot. Hotel']
    _hdr(ws2, 1, hdrs2)
    for i, m in enumerate(mesi):
        row = i + 2
        vals = [m['nome_mese']]
        tot_ts = 0.0
        for s in strutture_hotel:
            ts = m['per_struttura'].get(s, {}).get('tassa_soggiorno') or 0.0
            vals.append(ts)
            tot_ts += ts
        vals.append(tot_ts)
        _data_row(ws2, row, vals, alt=i % 2 == 1)
    tot_vals2 = ['TOTALE ANNO']
    tot_ts_anno = 0.0
    for s in strutture_hotel:
        ts = tot_anno['per_struttura'].get(s, {}).get('tassa_soggiorno') or 0.0
        tot_vals2.append(ts)
        tot_ts_anno += ts
    tot_vals2.append(tot_ts_anno)
    _tot_row(ws2, len(mesi) + 2, tot_vals2)
    _autowidth(ws2)

    # ── Foglio 3: Riepilogo di controllo ─────────────────────────────────────
    ws3 = wb.create_sheet('Riepilogo Controllo')
    hdrs3 = ['Mese', 'Corrispettivo', '+ Tassa Soggiorno', '= Totale Lordo']

    for c, val in enumerate(hdrs3, 1):
        cell = ws3.cell(row=1, column=c, value=val)
        cell.font = HDR_FONT
        cell.fill = TEAL_FILL if c in (2, 3) else HDR_FILL
        cell.alignment = Alignment(horizontal='center')

    ts_per_mese = {
        m['mese']: sum(
            m['per_struttura'].get(s, {}).get('tassa_soggiorno') or 0.0
            for s in strutture
        )
        for m in mesi
    }
    for i, m in enumerate(mesi):
        row = i + 2
        corr = m.get('totale_generale') or 0.0
        ts = ts_per_mese[m['mese']]
        _data_row(ws3, row, [m['nome_mese'], corr, ts, corr + ts], alt=i % 2 == 1)

    ts_anno = sum(
        tot_anno['per_struttura'].get(s, {}).get('tassa_soggiorno') or 0.0
        for s in strutture
    )
    corr_anno = tot_anno.get('totale_generale') or 0.0
    _tot_row(ws3, len(mesi) + 2, ['TOTALE ANNO', corr_anno, ts_anno, corr_anno + ts_anno])
    _autowidth(ws3)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f'corrispettivi_fatturati_{anno}.xlsx'
    return StreamingResponse(
        buf,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'},
    )


# ── RT Chiusure — Controllo RT ────────────────────────────────────────────────

# RT1 = DPH + CLB (un'unica cassa fiscale), RT2 = INT
RT_STRUTTURE: dict = {
    'RT1': ['DPH', 'CLB'],
    'RT2': ['INT'],
}


def _fmt_rt(r: RtChiusura) -> dict:
    return {
        'id': r.id,
        'data_chiusura': r.data_chiusura.isoformat(),
        'rt_code': r.rt_code,
        'totale_giorno': float(r.totale_giorno),
        'totale_10': float(r.totale_10) if r.totale_10 is not None else None,
        'totale_22': float(r.totale_22) if r.totale_22 is not None else None,
        'totale_ts': float(r.totale_ts) if r.totale_ts is not None else None,
        'totale_penali': float(r.totale_penali) if r.totale_penali is not None else None,
        'note': r.note,
        'created_at': r.created_at.isoformat() if r.created_at else None,
    }


@router.post("/rt-chiusure")
def upsert_rt_chiusura(
    body: dict,
    db: Session = Depends(get_db),
    utente=Depends(richiedi_admin),
):
    """Inserisce o aggiorna la chiusura RT per una data. Upsert su (data_chiusura, rt_code)."""
    rt_code = body.get('rt_code')
    if rt_code not in RT_STRUTTURE:
        raise HTTPException(400, f"rt_code non valido: usa {list(RT_STRUTTURE.keys())}")
    try:
        data = date.fromisoformat(body['data_chiusura'])
    except (KeyError, ValueError):
        raise HTTPException(400, "data_chiusura non valida (formato YYYY-MM-DD)")
    if body.get('totale_giorno') is None:
        raise HTTPException(400, "totale_giorno obbligatorio")

    def _dec(v):
        return Decimal(str(v)) if v is not None else None

    esistente = db.query(RtChiusura).filter(
        RtChiusura.data_chiusura == data,
        RtChiusura.rt_code == rt_code,
    ).first()

    if esistente:
        esistente.totale_giorno = _dec(body['totale_giorno'])
        esistente.totale_10 = _dec(body.get('totale_10'))
        esistente.totale_22 = _dec(body.get('totale_22'))
        esistente.totale_ts = _dec(body.get('totale_ts'))
        esistente.totale_penali = _dec(body.get('totale_penali'))
        esistente.note = body.get('note')
        esistente.updated_by = utente.id
        db.commit()
        db.refresh(esistente)
        return _fmt_rt(esistente)

    nuovo = RtChiusura(
        data_chiusura=data,
        rt_code=rt_code,
        totale_giorno=_dec(body['totale_giorno']),
        totale_10=_dec(body.get('totale_10')),
        totale_22=_dec(body.get('totale_22')),
        totale_ts=_dec(body.get('totale_ts')),
        totale_penali=_dec(body.get('totale_penali')),
        note=body.get('note'),
        created_by=utente.id,
    )
    db.add(nuovo)
    db.commit()
    db.refresh(nuovo)
    return _fmt_rt(nuovo)


@router.get("/rt-chiusure")
def lista_rt_chiusure(
    mese: int = Query(..., ge=1, le=12),
    anno: int = Query(..., ge=2020),
    db: Session = Depends(get_db),
    _utente=Depends(richiedi_utente_attivo),
):
    """Restituisce i giorni del mese con chiusure RT e il confronto vs scontrini PMS (lordi)."""
    from calendar import monthrange
    from datetime import date as date_t

    _, ultimo_giorno = monthrange(anno, mese)
    data_da = date_t(anno, mese, 1)
    data_a = date_t(anno, mese, ultimo_giorno)

    # Aggregazione scontrini PMS per data e struttura, breakdown per natura IVA.
    # Per la tassa soggiorno usiamo la COLONNA tassa_soggiorno (formato esteso) che
    # cattura sia la TS disaggregata dagli arrangiamenti sia i documenti standalone.
    # Fallback per formato base (colonna NULL): usa totale_lordo dei doc categoria='tassa_soggiorno'.
    # Per gli arrangiamenti, sottraiamo la parte TS già conteggiata nella colonna.
    rows_pms = db.execute(text("""
        SELECT
            data_documento,
            struttura_code,
            SUM(totale_lordo) FILTER (WHERE tipo = 'scontrino') AS totale,
            SUM(CASE
                WHEN tipo='scontrino' AND categoria='arrangiamenti' AND tassa_soggiorno IS NOT NULL
                    THEN totale_lordo - tassa_soggiorno
                WHEN tipo='scontrino' AND categoria='arrangiamenti'
                    THEN totale_lordo
                ELSE 0
            END) AS arr,
            SUM(totale_lordo) FILTER (WHERE tipo = 'scontrino' AND categoria = 'shop') AS shop,
            SUM(CASE
                WHEN tipo='scontrino' AND tassa_soggiorno IS NOT NULL
                    THEN tassa_soggiorno
                WHEN tipo='scontrino' AND categoria='tassa_soggiorno'
                    THEN totale_lordo
                ELSE 0
            END) AS ts,
            SUM(totale_lordo) FILTER (WHERE tipo = 'scontrino' AND categoria = 'penali') AS penali
        FROM corrispettivi_documenti
        WHERE data_documento BETWEEN :da AND :a
          AND struttura_code = ANY(:strutture)
        GROUP BY data_documento, struttura_code
    """), {'da': data_da, 'a': data_a, 'strutture': STRUTTURE_HOTEL}).fetchall()

    # Indice: data_iso → struttura_code → {totale, arr, shop, ts, penali}
    pms_idx: dict = {}
    for row in rows_pms:
        d = row.data_documento.isoformat()
        pms_idx.setdefault(d, {})[row.struttura_code] = {
            'totale': float(row.totale or 0),
            'arr':    float(row.arr or 0),
            'shop':   float(row.shop or 0),
            'ts':     float(row.ts or 0),
            'penali': float(row.penali or 0),
        }

    # Chiusure RT del mese
    rt_rows = db.query(RtChiusura).filter(
        RtChiusura.data_chiusura >= data_da,
        RtChiusura.data_chiusura <= data_a,
    ).all()

    rt_idx: dict = {}
    for r in rt_rows:
        rt_idx.setdefault(r.data_chiusura.isoformat(), {})[r.rt_code] = r

    def _pms_agg(data_iso: str, strutture: list) -> dict:
        totale = arr = shop = ts = penali = 0.0
        for s in strutture:
            v = pms_idx.get(data_iso, {}).get(s, {})
            totale += v.get('totale', 0.0)
            arr    += v.get('arr', 0.0)
            shop   += v.get('shop', 0.0)
            ts     += v.get('ts', 0.0)
            penali += v.get('penali', 0.0)
        return {'totale': totale, 'arr': arr, 'shop': shop, 'ts': ts, 'penali': penali}

    def _confronta(rt: Optional[RtChiusura], pms: dict) -> dict:
        if rt is None:
            return {'rt': None, 'pms': pms, 'delta': None, 'breakdown': None}
        rt_tot = float(rt.totale_giorno)
        delta = round(rt_tot - pms['totale'], 2)
        # Breakdown analitico solo se tutti i 4 campi natura IVA sono compilati
        breakdown = None
        if all(v is not None for v in [rt.totale_10, rt.totale_22, rt.totale_ts, rt.totale_penali]):
            breakdown = {
                '10':     {'rt': float(rt.totale_10),     'pms': pms['arr'],    'delta': round(float(rt.totale_10)     - pms['arr'],    2)},
                '22':     {'rt': float(rt.totale_22),     'pms': pms['shop'],   'delta': round(float(rt.totale_22)     - pms['shop'],   2)},
                'ts':     {'rt': float(rt.totale_ts),     'pms': pms['ts'],     'delta': round(float(rt.totale_ts)     - pms['ts'],     2)},
                'penali': {'rt': float(rt.totale_penali), 'pms': pms['penali'], 'delta': round(float(rt.totale_penali) - pms['penali'], 2)},
            }
        return {
            'rt': {
                'id':            rt.id,
                'totale_giorno': rt_tot,
                'totale_10':     float(rt.totale_10)     if rt.totale_10     is not None else None,
                'totale_22':     float(rt.totale_22)     if rt.totale_22     is not None else None,
                'totale_ts':     float(rt.totale_ts)     if rt.totale_ts     is not None else None,
                'totale_penali': float(rt.totale_penali) if rt.totale_penali is not None else None,
                'note':          rt.note,
            },
            'pms': pms,
            'delta': delta,
            'breakdown': breakdown,
        }

    giorni = []
    for g in range(1, ultimo_giorno + 1):
        d = date_t(anno, mese, g)
        d_iso = d.isoformat()
        pms1 = _pms_agg(d_iso, RT_STRUTTURE['RT1'])
        pms2 = _pms_agg(d_iso, RT_STRUTTURE['RT2'])
        rt1 = rt_idx.get(d_iso, {}).get('RT1')
        rt2 = rt_idx.get(d_iso, {}).get('RT2')
        # Includi il giorno solo se c'è almeno un dato (PMS o RT)
        if pms1['totale'] == 0 and pms2['totale'] == 0 and not rt1 and not rt2:
            continue
        giorni.append({
            'data': d_iso,
            'rt1': _confronta(rt1, pms1),
            'rt2': _confronta(rt2, pms2),
        })

    n_differenze = sum(
        1 for g in giorni
        if (g['rt1']['delta'] is not None and abs(g['rt1']['delta']) > 0.01)
        or (g['rt2']['delta'] is not None and abs(g['rt2']['delta']) > 0.01)
    )

    return {'mese': mese, 'anno': anno, 'giorni': giorni, 'n_differenze': n_differenze}


@router.delete("/rt-chiusure/{chiusura_id}")
def elimina_rt_chiusura(
    chiusura_id: int,
    db: Session = Depends(get_db),
    _utente=Depends(richiedi_admin),
):
    """Elimina una chiusura RT."""
    rt = db.query(RtChiusura).filter(RtChiusura.id == chiusura_id).first()
    if not rt:
        raise HTTPException(404, "Chiusura RT non trovata")
    db.delete(rt)
    db.commit()
    return {'ok': True}
