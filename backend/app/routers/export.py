"""Router FastAPI — export dati in formato Excel (.xlsx), CSV e PDF."""

import csv
import io
from collections import defaultdict
from datetime import date
from typing import Dict, List, Optional

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse

from app.auth import richiedi_utente_attivo
from sqlalchemy import select
from sqlalchemy.orm import Session

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet

from app.database import get_db
from app.models.revenue import DailyRevenue, Hotel
from app.routers.dashboard import _carica_righe, _db_a_riga
from app.services.file_parser import RigaRevenue
from app.services.kpi_calculator import calcola_kpi
from app.services.weekly_aggregator import AggregatoSettimanale, aggrega_settimane
from app.services.group_aggregator import aggrega_gruppo_settimanale

from app.utils.locale_it import GIORNI_IT

router = APIRouter(prefix="/export", tags=["export"], dependencies=[Depends(richiedi_utente_attivo)])
_BLU_INTESTAZIONE = "1a56db"
_GRIGIO_RIGA      = "f1f5f9"

# Formato numerico per celle Excel
_FMT_EURO = '#,##0.00 "€"'
_FMT_PERC = '0.0"%"'
_FMT_INT  = '#,##0'


# ---------------------------------------------------------------------------
# Endpoint hotel — settimanale
# ---------------------------------------------------------------------------

@router.get("/hotel/{hotel_code}/settimanale")
def export_hotel_settimanale(
    hotel_code: str,
    da: Optional[date] = Query(None),
    a: Optional[date] = Query(None),
    snapshot: Optional[date] = Query(None),
    formato: str = Query("xlsx", pattern="^(xlsx|csv|pdf)$"),
    db: Session = Depends(get_db),
):
    """Esporta gli aggregati settimanali di un hotel in Excel, CSV o PDF."""
    hotel_code = hotel_code.upper()
    righe = _carica_righe(db, hotel_code=hotel_code, snapshot_date=snapshot, da=da, a=a)
    if not righe:
        raise HTTPException(status_code=404, detail="Nessun dato nel periodo selezionato.")
    settimane = aggrega_settimane(righe)
    nome = f"{hotel_code}_settimanale"
    return _risposta(formato, nome, _xlsx_hotel_sett, _csv_hotel_sett, _pdf_hotel_sett, hotel_code, settimane)


# ---------------------------------------------------------------------------
# Endpoint hotel — giornaliero
# ---------------------------------------------------------------------------

@router.get("/hotel/{hotel_code}/giornaliero")
def export_hotel_giornaliero(
    hotel_code: str,
    da: Optional[date] = Query(None),
    a: Optional[date] = Query(None),
    snapshot: Optional[date] = Query(None),
    formato: str = Query("xlsx", pattern="^(xlsx|csv|pdf)$"),
    db: Session = Depends(get_db),
):
    """Esporta i dati giornalieri di un hotel in Excel, CSV o PDF."""
    hotel_code = hotel_code.upper()
    righe = _carica_righe(db, hotel_code=hotel_code, snapshot_date=snapshot, da=da, a=a)
    if not righe:
        raise HTTPException(status_code=404, detail="Nessun dato nel periodo selezionato.")
    nome = f"{hotel_code}_giornaliero"
    return _risposta(formato, nome, _xlsx_hotel_giorn, _csv_hotel_giorn, _pdf_hotel_giorn, hotel_code, righe)


# ---------------------------------------------------------------------------
# Endpoint gruppo — settimanale + dettaglio hotel
# ---------------------------------------------------------------------------

@router.get("/gruppo")
def export_gruppo(
    da: Optional[date] = Query(None),
    a: Optional[date] = Query(None),
    formato: str = Query("xlsx", pattern="^(xlsx|csv|pdf)$"),
    db: Session = Depends(get_db),
):
    """Esporta aggregati settimanali di gruppo e dettaglio per hotel in Excel, CSV o PDF."""
    q = select(DailyRevenue).order_by(DailyRevenue.hotel_code, DailyRevenue.data)
    if da:
        q = q.where(DailyRevenue.data >= da)
    if a:
        q = q.where(DailyRevenue.data <= a)
    rows = db.execute(q).scalars().all()
    if not rows:
        raise HTTPException(status_code=404, detail="Nessun dato nel periodo selezionato.")

    per_hotel: Dict[str, List[RigaRevenue]] = defaultdict(list)
    for row in rows:
        per_hotel[row.hotel_code].append(_db_a_riga(row))

    hotel_nomi = {h.code: h.name for h in db.query(Hotel).all()}
    settimane  = aggrega_gruppo_settimanale(per_hotel)
    contributi = _calcola_contributi(per_hotel, hotel_nomi)
    return _risposta(formato, "gruppo_settimanale", _xlsx_gruppo, _csv_gruppo, _pdf_gruppo,
                     "GRUPPO", settimane, contributi)


# ---------------------------------------------------------------------------
# Dispatcher
# ---------------------------------------------------------------------------

def _risposta(formato, nome, fn_xlsx, fn_csv, fn_pdf, *args):
    if formato == "xlsx":
        buf = fn_xlsx(*args)
        mt = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ext = "xlsx"
    elif formato == "csv":
        buf = fn_csv(*args)
        mt = "text/csv; charset=utf-8-sig"
        ext = "csv"
    else:
        buf = fn_pdf(*args)
        mt = "application/pdf"
        ext = "pdf"
    return StreamingResponse(
        buf,
        media_type=mt,
        headers={"Content-Disposition": f'attachment; filename="{nome}.{ext}"'},
    )


# ---------------------------------------------------------------------------
# Excel — hotel settimanale (16 colonne)
# ---------------------------------------------------------------------------

_INT_SETT = [
    "Settimana", "Giorni", "Cam. Vendute", "Cam. Disponibili",
    "Occup. %", "ADR (€)", "RevPAR (€)", "TRevPAR (€)", "RMC (€)",
    "Rev. Camere (€)", "Rev. F&B (€)", "Rev. Extra (€)", "Rev. Totale (€)",
    "Inc. Rooms (%)", "Inc. F&B (%)", "Inc. Extra (%)",
]


def _riga_sett(s: AggregatoSettimanale) -> list:
    return [
        f"{s.week_start.strftime('%d/%m/%Y')}–{s.week_end.strftime('%d/%m/%Y')}",
        s.giorni,
        s.rooms_sold,
        s.rooms_available,
        s.kpi.occupancy,
        s.kpi.adr,
        s.kpi.revpar,
        s.kpi.trevpar,
        s.kpi.rmc,
        round(s.revenue_rooms, 2),
        round(s.revenue_fnb, 2),
        round(s.revenue_extra, 2),
        round(s.revenue_total, 2),
        s.kpi.inc_rooms,
        s.kpi.inc_fnb,
        s.kpi.inc_extra,
    ]


def _xlsx_hotel_sett(hotel_code: str, settimane: List[AggregatoSettimanale]) -> io.BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Settimanale"
    _scrivi_intestazione(ws, _INT_SETT)

    for i, s in enumerate(settimane, start=2):
        _scrivi_riga_sett(ws, i, _riga_sett(s))

    # Riga totale stagione in grassetto
    tot = _totale_sett(settimane)
    riga_tot = len(settimane) + 2
    _scrivi_riga_sett(ws, riga_tot, tot, bold=True)

    _auto_width(ws)
    return _to_buf(wb)


def _scrivi_riga_sett(ws, row_num: int, valori: list, bold: bool = False) -> None:
    """Scrive una riga con formattazione numero appropriata per ogni colonna."""
    # Formati per le 16 colonne: stringa, int, int, int, %, €, €, €, €, €, €, €, €, %, %, %
    formati = [
        None, _FMT_INT, _FMT_INT, _FMT_INT,
        _FMT_PERC, _FMT_EURO, _FMT_EURO, _FMT_EURO, _FMT_EURO,
        _FMT_EURO, _FMT_EURO, _FMT_EURO, _FMT_EURO,
        _FMT_PERC, _FMT_PERC, _FMT_PERC,
    ]
    for j, (val, fmt) in enumerate(zip(valori, formati), start=1):
        cell = ws.cell(row=row_num, column=j, value=val)
        if fmt:
            cell.number_format = fmt
        if bold:
            cell.font = Font(bold=True)
        elif row_num % 2 == 0:
            cell.fill = PatternFill("solid", fgColor=_GRIGIO_RIGA)


def _totale_sett(settimane: List[AggregatoSettimanale]) -> list:
    """Calcola la riga aggregata totale stagione."""
    rs = sum(s.rooms_sold for s in settimane)
    ra = sum(s.rooms_available for s in settimane)
    rr = sum(s.revenue_rooms for s in settimane)
    rf = sum(s.revenue_fnb for s in settimane)
    rx = sum(s.revenue_extra for s in settimane)
    rt = sum(s.revenue_total for s in settimane)
    kpi = calcola_kpi(rs, ra, rr, rf, rx, rt)
    return [
        "TOTALE STAGIONE",
        sum(s.giorni for s in settimane),
        rs, ra,
        kpi.occupancy, kpi.adr, kpi.revpar, kpi.trevpar, kpi.rmc,
        round(rr, 2), round(rf, 2), round(rx, 2), round(rt, 2),
        kpi.inc_rooms, kpi.inc_fnb, kpi.inc_extra,
    ]


# ---------------------------------------------------------------------------
# Excel — hotel giornaliero (13 colonne)
# ---------------------------------------------------------------------------

_INT_GIORN = [
    "Data", "Giorno", "Cam. Vendute", "PAX",
    "Occup. %", "ADR (€)", "RMC (€)", "RevPAR (€)", "TRevPAR (€)",
    "Rev. Camere (€)", "Rev. F&B (€)", "Rev. Extra (€)", "Rev. Totale (€)",
]
_FMT_GIORN = [
    None, None, _FMT_INT, _FMT_INT,
    _FMT_PERC, _FMT_EURO, _FMT_EURO, _FMT_EURO, _FMT_EURO,
    _FMT_EURO, _FMT_EURO, _FMT_EURO, _FMT_EURO,
]


def _xlsx_hotel_giorn(hotel_code: str, righe: List[RigaRevenue]) -> io.BytesIO:
    from app.services.kpi_calculator import kpi_da_riga
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Giornaliero"
    _scrivi_intestazione(ws, _INT_GIORN)

    for i, r in enumerate(righe, start=2):
        kpi = kpi_da_riga(r)
        valori = [
            r.data.strftime("%d/%m/%Y"),
            GIORNI_IT[r.data.weekday()],
            r.rooms_sold,
            r.pax,
            kpi.occupancy,
            kpi.adr,
            kpi.rmc,
            kpi.revpar,
            kpi.trevpar,
            round(r.revenue_rooms, 2),
            round(r.revenue_fnb, 2),
            round(r.revenue_extra, 2),
            round(r.revenue_total, 2),
        ]
        for j, (val, fmt) in enumerate(zip(valori, _FMT_GIORN), start=1):
            cell = ws.cell(row=i, column=j, value=val)
            if fmt:
                cell.number_format = fmt
            if i % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=_GRIGIO_RIGA)

    _auto_width(ws)
    return _to_buf(wb)


# ---------------------------------------------------------------------------
# Excel — gruppo settimanale (2 fogli)
# ---------------------------------------------------------------------------

_INT_GRUPPO_SETT = [
    "Settimana", "Hotel attivi", "Cam. Vendute", "Cam. Disponibili",
    "Occup. %", "ADR (€)", "RevPAR (€)", "TRevPAR (€)", "RMC (€)",
    "Rev. Camere (€)", "Rev. F&B (€)", "Rev. Extra (€)", "Rev. Totale (€)",
    "Inc. Rooms (%)", "Inc. F&B (%)", "Inc. Extra (%)",
]
_FMT_GRUPPO_SETT = [
    None, None, _FMT_INT, _FMT_INT,
    _FMT_PERC, _FMT_EURO, _FMT_EURO, _FMT_EURO, _FMT_EURO,
    _FMT_EURO, _FMT_EURO, _FMT_EURO, _FMT_EURO,
    _FMT_PERC, _FMT_PERC, _FMT_PERC,
]

_INT_GRUPPO_HOTEL = [
    "Codice", "Hotel", "Cam. Vendute", "Cam. Disponibili",
    "Occup. %", "ADR (€)", "RevPAR (€)",
    "Rev. Camere (€)", "Rev. F&B (€)", "Rev. Extra (€)", "Rev. Totale (€)",
    "% Gruppo",
]
_FMT_GRUPPO_HOTEL = [
    None, None, _FMT_INT, _FMT_INT,
    _FMT_PERC, _FMT_EURO, _FMT_EURO,
    _FMT_EURO, _FMT_EURO, _FMT_EURO, _FMT_EURO,
    _FMT_PERC,
]


def _xlsx_gruppo(label: str, settimane, contributi: list) -> io.BytesIO:
    wb = openpyxl.Workbook()

    # Foglio 1: aggregati settimanali
    ws1 = wb.active
    ws1.title = "Aggregati settimanali"
    _scrivi_intestazione(ws1, _INT_GRUPPO_SETT)

    for i, s in enumerate(settimane, start=2):
        valori = [
            f"{s.week_start.strftime('%d/%m/%Y')}–{s.week_end.strftime('%d/%m/%Y')}",
            ", ".join(s.hotel_codes),
            s.rooms_sold, s.rooms_available,
            s.kpi.occupancy, s.kpi.adr, s.kpi.revpar, s.kpi.trevpar, s.kpi.rmc,
            round(s.revenue_rooms, 2), round(s.revenue_fnb, 2),
            round(s.revenue_extra, 2), round(s.revenue_total, 2),
            s.kpi.inc_rooms, s.kpi.inc_fnb, s.kpi.inc_extra,
        ]
        for j, (val, fmt) in enumerate(zip(valori, _FMT_GRUPPO_SETT), start=1):
            cell = ws1.cell(row=i, column=j, value=val)
            if fmt:
                cell.number_format = fmt
            if i % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=_GRIGIO_RIGA)

    # Riga totale
    tot = _totale_gruppo_sett(settimane)
    riga_tot = len(settimane) + 2
    for j, (val, fmt) in enumerate(zip(tot, _FMT_GRUPPO_SETT), start=1):
        cell = ws1.cell(row=riga_tot, column=j, value=val)
        if fmt:
            cell.number_format = fmt
        cell.font = Font(bold=True)
    _auto_width(ws1)

    # Foglio 2: dettaglio per hotel
    ws2 = wb.create_sheet("Dettaglio hotel")
    _scrivi_intestazione(ws2, _INT_GRUPPO_HOTEL)
    for i, c in enumerate(contributi, start=2):
        valori = [
            c["hotel_code"], c["hotel_name"],
            c["rooms_sold"], c["rooms_available"],
            c["occupancy"], c["adr"], c["revpar"],
            c["revenue_rooms"], c["revenue_fnb"], c["revenue_extra"], c["revenue_total"],
            c["perc_revenue"],
        ]
        for j, (val, fmt) in enumerate(zip(valori, _FMT_GRUPPO_HOTEL), start=1):
            cell = ws2.cell(row=i, column=j, value=val)
            if fmt:
                cell.number_format = fmt
            if i % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=_GRIGIO_RIGA)
    _auto_width(ws2)

    return _to_buf(wb)


def _totale_gruppo_sett(settimane) -> list:
    rs = sum(s.rooms_sold for s in settimane)
    ra = sum(s.rooms_available for s in settimane)
    rr = sum(s.revenue_rooms for s in settimane)
    rf = sum(s.revenue_fnb for s in settimane)
    rx = sum(s.revenue_extra for s in settimane)
    rt = sum(s.revenue_total for s in settimane)
    hotel = sorted({h for s in settimane for h in s.hotel_codes})
    kpi = calcola_kpi(rs, ra, rr, rf, rx, rt)
    return [
        "TOTALE STAGIONE", ", ".join(hotel),
        rs, ra,
        kpi.occupancy, kpi.adr, kpi.revpar, kpi.trevpar, kpi.rmc,
        round(rr, 2), round(rf, 2), round(rx, 2), round(rt, 2),
        kpi.inc_rooms, kpi.inc_fnb, kpi.inc_extra,
    ]


# ---------------------------------------------------------------------------
# CSV — hotel settimanale
# ---------------------------------------------------------------------------

def _csv_hotel_sett(hotel_code: str, settimane: List[AggregatoSettimanale]) -> io.BytesIO:
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(_INT_SETT)
    for s in settimane:
        w.writerow([_sv(v) for v in _riga_sett(s)])
    w.writerow([_sv(v) for v in _totale_sett(settimane)])
    return io.BytesIO(b"\xef\xbb\xbf" + buf.getvalue().encode("utf-8"))


# ---------------------------------------------------------------------------
# CSV — hotel giornaliero
# ---------------------------------------------------------------------------

def _csv_hotel_giorn(hotel_code: str, righe: List[RigaRevenue]) -> io.BytesIO:
    from app.services.kpi_calculator import kpi_da_riga
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(_INT_GIORN)
    for r in righe:
        kpi = kpi_da_riga(r)
        w.writerow([
            r.data.isoformat(), GIORNI_IT[r.data.weekday()],
            r.rooms_sold, r.pax,
            _sv(kpi.occupancy), _sv(kpi.adr), _sv(kpi.rmc), _sv(kpi.revpar), _sv(kpi.trevpar),
            round(r.revenue_rooms, 2), round(r.revenue_fnb, 2),
            round(r.revenue_extra, 2), round(r.revenue_total, 2),
        ])
    return io.BytesIO(b"\xef\xbb\xbf" + buf.getvalue().encode("utf-8"))


# ---------------------------------------------------------------------------
# CSV — gruppo
# ---------------------------------------------------------------------------

def _csv_gruppo(label: str, settimane, contributi: list) -> io.BytesIO:
    buf = io.StringIO()
    w = csv.writer(buf)

    # Sezione 1: aggregati settimanali
    w.writerow(["=== AGGREGATI SETTIMANALI GRUPPO ==="])
    w.writerow(_INT_GRUPPO_SETT)
    for s in settimane:
        w.writerow([
            f"{s.week_start.isoformat()}–{s.week_end.isoformat()}",
            "|".join(s.hotel_codes),
            s.rooms_sold, s.rooms_available,
            _sv(s.kpi.occupancy), _sv(s.kpi.adr), _sv(s.kpi.revpar),
            _sv(s.kpi.trevpar), _sv(s.kpi.rmc),
            round(s.revenue_rooms, 2), round(s.revenue_fnb, 2),
            round(s.revenue_extra, 2), round(s.revenue_total, 2),
            _sv(s.kpi.inc_rooms), _sv(s.kpi.inc_fnb), _sv(s.kpi.inc_extra),
        ])
    tot = _totale_gruppo_sett(settimane)
    w.writerow([_sv(v) for v in tot])

    # Sezione 2: dettaglio per hotel
    w.writerow([])
    w.writerow(["=== DETTAGLIO PER HOTEL ==="])
    w.writerow(_INT_GRUPPO_HOTEL)
    for c in contributi:
        w.writerow([
            c["hotel_code"], c["hotel_name"],
            c["rooms_sold"], c["rooms_available"],
            _sv(c["occupancy"]), _sv(c["adr"]), _sv(c["revpar"]),
            c["revenue_rooms"], c["revenue_fnb"], c["revenue_extra"], c["revenue_total"],
            _sv(c["perc_revenue"]),
        ])
    return io.BytesIO(b"\xef\xbb\xbf" + buf.getvalue().encode("utf-8"))


# ---------------------------------------------------------------------------
# PDF — hotel settimanale (16 colonne abbreviate)
# ---------------------------------------------------------------------------

_PDF_INT_SETT = [
    "Settimana", "Gg.", "C.V.", "C.D.",
    "Occ%", "ADR", "RevPAR", "TRevPAR", "RMC",
    "R.Cam.", "R.F&B", "R.Extra", "R.Tot.",
    "I.Rm%", "I.FB%", "I.Ex%",
]
_PDF_CW_SETT = [88, 22, 30, 30, 36, 42, 44, 48, 42, 48, 42, 40, 48, 36, 36, 36]


def _pdf_hotel_sett(hotel_code: str, settimane: List[AggregatoSettimanale]) -> io.BytesIO:
    titolo = f"{hotel_code} — Aggregati settimanali"
    righe_pdf = []
    for s in settimane:
        righe_pdf.append([
            f"{s.week_start.strftime('%d/%m')}–{s.week_end.strftime('%d/%m/%Y')}",
            str(s.giorni), str(s.rooms_sold), str(s.rooms_available),
            _pv(s.kpi.occupancy, 1, "%"), _pv(s.kpi.adr, 2, "€"),
            _pv(s.kpi.revpar, 2, "€"), _pv(s.kpi.trevpar, 2, "€"), _pv(s.kpi.rmc, 2, "€"),
            _pv(s.revenue_rooms, 0, "€"), _pv(s.revenue_fnb, 0, "€"),
            _pv(s.revenue_extra, 0, "€"), _pv(s.revenue_total, 0, "€"),
            _pv(s.kpi.inc_rooms, 1, "%"), _pv(s.kpi.inc_fnb, 1, "%"), _pv(s.kpi.inc_extra, 1, "%"),
        ])
    tot = _totale_sett(settimane)
    righe_pdf.append([
        "TOTALE", str(tot[1]), str(tot[2]), str(tot[3]),
        _pv(tot[4], 1, "%"), _pv(tot[5], 2, "€"),
        _pv(tot[6], 2, "€"), _pv(tot[7], 2, "€"), _pv(tot[8], 2, "€"),
        _pv(tot[9], 0, "€"), _pv(tot[10], 0, "€"),
        _pv(tot[11], 0, "€"), _pv(tot[12], 0, "€"),
        _pv(tot[13], 1, "%"), _pv(tot[14], 1, "%"), _pv(tot[15], 1, "%"),
    ])
    n = len(righe_pdf)
    extra = [("FONTNAME", (0, n), (-1, n), "Helvetica-Bold")]
    return _costruisci_pdf(titolo, _PDF_INT_SETT, righe_pdf, _PDF_CW_SETT, extra_stili=extra)


# ---------------------------------------------------------------------------
# PDF — hotel giornaliero (13 colonne abbreviate)
# ---------------------------------------------------------------------------

_PDF_INT_GIORN = [
    "Data", "Giorno", "C.V.", "PAX",
    "Occ%", "ADR", "RMC", "RevPAR", "TRevPAR",
    "R.Cam.", "R.F&B", "R.Extra", "R.Tot.",
]
_PDF_CW_GIORN = [50, 34, 28, 26, 34, 42, 42, 44, 48, 48, 42, 38, 48]


def _pdf_hotel_giorn(hotel_code: str, righe: List[RigaRevenue]) -> io.BytesIO:
    from app.services.kpi_calculator import kpi_da_riga
    titolo = f"{hotel_code} — Dati giornalieri"
    righe_pdf = []
    for r in righe:
        kpi = kpi_da_riga(r)
        righe_pdf.append([
            r.data.strftime("%d/%m/%Y"), GIORNI_IT[r.data.weekday()],
            str(r.rooms_sold), str(r.pax),
            _pv(kpi.occupancy, 1, "%"), _pv(kpi.adr, 2, "€"),
            _pv(kpi.rmc, 2, "€"), _pv(kpi.revpar, 2, "€"), _pv(kpi.trevpar, 2, "€"),
            _pv(r.revenue_rooms, 0, "€"), _pv(r.revenue_fnb, 0, "€"),
            _pv(r.revenue_extra, 0, "€"), _pv(r.revenue_total, 0, "€"),
        ])
    return _costruisci_pdf(titolo, _PDF_INT_GIORN, righe_pdf, _PDF_CW_GIORN)


# ---------------------------------------------------------------------------
# PDF — gruppo (2 tabelle)
# ---------------------------------------------------------------------------

_PDF_INT_GRUPPO_SETT = [
    "Settimana", "Hotel", "C.V.", "C.D.",
    "Occ%", "ADR", "RevPAR", "TRevPAR", "RMC",
    "R.Cam.", "R.F&B", "R.Extra", "R.Tot.",
    "I.Rm%", "I.FB%", "I.Ex%",
]
_PDF_CW_GRUPPO_SETT = [80, 36, 30, 30, 36, 42, 44, 48, 42, 48, 42, 40, 48, 36, 36, 36]

_PDF_INT_GRUPPO_HOTEL = [
    "Cod.", "Hotel", "C.V.", "C.D.",
    "Occ%", "ADR", "RevPAR",
    "R.Cam.", "R.F&B", "R.Extra", "R.Tot.", "% Gr.",
]
_PDF_CW_GRUPPO_HOTEL = [32, 90, 36, 36, 36, 48, 50, 62, 55, 50, 62, 40]


def _pdf_gruppo(label: str, settimane, contributi: list) -> io.BytesIO:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        leftMargin=1.5 * 28.35, rightMargin=1.5 * 28.35,
        topMargin=1.5 * 28.35, bottomMargin=1.5 * 28.35,
    )
    stili = getSampleStyleSheet()
    elementi = [Paragraph("Gruppo — Aggregati settimanali", stili["Heading2"]), Spacer(1, 8)]

    righe_pdf = []
    for s in settimane:
        righe_pdf.append([
            f"{s.week_start.strftime('%d/%m')}–{s.week_end.strftime('%d/%m/%Y')}",
            ",".join(s.hotel_codes),
            str(s.rooms_sold), str(s.rooms_available),
            _pv(s.kpi.occupancy, 1, "%"), _pv(s.kpi.adr, 2, "€"),
            _pv(s.kpi.revpar, 2, "€"), _pv(s.kpi.trevpar, 2, "€"), _pv(s.kpi.rmc, 2, "€"),
            _pv(s.revenue_rooms, 0, "€"), _pv(s.revenue_fnb, 0, "€"),
            _pv(s.revenue_extra, 0, "€"), _pv(s.revenue_total, 0, "€"),
            _pv(s.kpi.inc_rooms, 1, "%"), _pv(s.kpi.inc_fnb, 1, "%"), _pv(s.kpi.inc_extra, 1, "%"),
        ])
    tot = _totale_gruppo_sett(settimane)
    righe_pdf.append([
        "TOTALE", "", str(tot[2]), str(tot[3]),
        _pv(tot[4], 1, "%"), _pv(tot[5], 2, "€"),
        _pv(tot[6], 2, "€"), _pv(tot[7], 2, "€"), _pv(tot[8], 2, "€"),
        _pv(tot[9], 0, "€"), _pv(tot[10], 0, "€"),
        _pv(tot[11], 0, "€"), _pv(tot[12], 0, "€"),
        _pv(tot[13], 1, "%"), _pv(tot[14], 1, "%"), _pv(tot[15], 1, "%"),
    ])
    n = len(righe_pdf)
    extra = [("FONTNAME", (0, n), (-1, n), "Helvetica-Bold")]
    t1 = _crea_tabella(_PDF_INT_GRUPPO_SETT, righe_pdf, _PDF_CW_GRUPPO_SETT, extra_stili=extra)
    elementi.append(t1)

    if contributi:
        elementi += [Spacer(1, 18), Paragraph("Dettaglio per hotel", stili["Heading2"]), Spacer(1, 8)]
        righe_hotel = []
        for c in contributi:
            righe_hotel.append([
                c["hotel_code"], c["hotel_name"],
                str(c["rooms_sold"]), str(c["rooms_available"]),
                _pv(c["occupancy"], 1, "%"), _pv(c["adr"], 2, "€"), _pv(c["revpar"], 2, "€"),
                _pv(c["revenue_rooms"], 0, "€"), _pv(c["revenue_fnb"], 0, "€"),
                _pv(c["revenue_extra"], 0, "€"), _pv(c["revenue_total"], 0, "€"),
                _pv(c["perc_revenue"], 1, "%"),
            ])
        t2 = _crea_tabella(_PDF_INT_GRUPPO_HOTEL, righe_hotel, _PDF_CW_GRUPPO_HOTEL)
        elementi.append(t2)

    doc.build(elementi)
    buf.seek(0)
    return buf


# ---------------------------------------------------------------------------
# Costruzione PDF generica
# ---------------------------------------------------------------------------

def _crea_tabella(intestazioni, righe, col_widths=None, extra_stili=None) -> Table:
    dati = [intestazioni] + righe
    t = Table(dati, colWidths=col_widths, repeatRows=1)
    stile = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1a56db")),
        ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
        ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",   (0, 0), (-1, -1), 8),
        ("ALIGN",      (2, 0), (-1, -1), "RIGHT"),
        ("ALIGN",      (0, 0), (1, -1), "LEFT"),
        ("GRID",       (0, 0), (-1, -1), 0.4, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f1f5f9")]),
        ("TOPPADDING",    (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]
    if extra_stili:
        stile.extend(extra_stili)
    t.setStyle(TableStyle(stile))
    return t


def _costruisci_pdf(titolo: str, intestazioni: list, righe: list,
                    col_widths=None, extra_stili=None) -> io.BytesIO:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        leftMargin=1.5 * 28.35, rightMargin=1.5 * 28.35,
        topMargin=1.5 * 28.35, bottomMargin=1.5 * 28.35,
    )
    stili = getSampleStyleSheet()
    elementi = [
        Paragraph(titolo, stili["Heading2"]),
        Spacer(1, 8),
        _crea_tabella(intestazioni, righe, col_widths, extra_stili),
    ]
    doc.build(elementi)
    buf.seek(0)
    return buf


# ---------------------------------------------------------------------------
# Utility condivise: _carica_righe/_db_a_riga importate da dashboard.py
# (unica fonte di verità per la lettura di daily_revenue, vedi CLAUDE.md)
# ---------------------------------------------------------------------------


def _calcola_contributi(per_hotel: dict, hotel_nomi: dict) -> list:
    """Calcola i totali per hotel (per dettaglio export gruppo)."""
    tot_revenue = sum(sum(r.revenue_total for r in righe) for righe in per_hotel.values())
    risultati = []
    for codice in sorted(per_hotel.keys()):
        righe = per_hotel[codice]
        rs = sum(r.rooms_sold for r in righe)
        ra = sum(r.rooms_available for r in righe)
        rr = sum(r.revenue_rooms for r in righe)
        rf = sum(r.revenue_fnb for r in righe)
        rx = sum(r.revenue_extra for r in righe)
        rt = sum(r.revenue_total for r in righe)
        kpi = calcola_kpi(rs, ra, rr, rf, rx, rt)
        risultati.append({
            "hotel_code": codice,
            "hotel_name": hotel_nomi.get(codice, codice),
            "rooms_sold": rs,
            "rooms_available": ra,
            "revenue_rooms": round(rr, 2),
            "revenue_fnb": round(rf, 2),
            "revenue_extra": round(rx, 2),
            "revenue_total": round(rt, 2),
            "occupancy": round(kpi.occupancy, 2) if kpi.occupancy is not None else None,
            "adr":       round(kpi.adr, 2)       if kpi.adr is not None       else None,
            "revpar":    round(kpi.revpar, 2)     if kpi.revpar is not None    else None,
            "perc_revenue": round(rt / tot_revenue * 100, 2) if tot_revenue else None,
        })
    return risultati


def _v(val: Optional[float], decimali: int = 2) -> str:
    """Stringa formattata per PDF (vecchio stile, mantenuto per compatibilità)."""
    if val is None:
        return "—"
    return f"{val:.{decimali}f}"


def _sv(val) -> str:
    """Stringa CSV: None → stringa vuota, float → 2 decimali."""
    if val is None:
        return ""
    if isinstance(val, float):
        return f"{val:.2f}"
    return str(val)


def _pv(val, decimali: int, suffisso: str = "") -> str:
    """Stringa PDF con suffisso (€ o %)."""
    if val is None:
        return "—"
    if suffisso == "€":
        return f"€{val:,.{decimali}f}"
    return f"{val:.{decimali}f}{suffisso}"


def _scrivi_intestazione(ws, intestazioni: list) -> None:
    fill = PatternFill("solid", fgColor=_BLU_INTESTAZIONE)
    font = Font(bold=True, color="FFFFFF")
    for j, testo in enumerate(intestazioni, start=1):
        cell = ws.cell(row=1, column=j, value=testo)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")


def _auto_width(ws) -> None:
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 3, 35)


def _to_buf(wb: openpyxl.Workbook) -> io.BytesIO:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
