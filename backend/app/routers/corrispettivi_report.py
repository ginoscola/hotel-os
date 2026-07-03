"""Sotto-router Corrispettivi — report aggregati, check, export ed export Excel.

Endpoint (montati sotto /corrispettivi dall'aggregatore corrispettivi.py):
  GET    /report/giornaliero    → aggregato per giorno e struttura
  GET    /report/mensile        → aggregato per mese e struttura
  GET    /check                 → totali per struttura
  GET    /report/fatturati      → riepilogo fatturati per mese/struttura
  GET    /report/pagamenti      → riepilogo per tipo di pagamento
  GET    /admin/test-stats      → conteggio record is_test
  DELETE /admin/test-data       → cancella tutti i record is_test
  GET    /export/fatturati      → export Excel riepilogo fatturati
"""
from datetime import date
from typing import List, Optional, Set

from fastapi import APIRouter, Depends, HTTPException, Query
from sqlalchemy import func
from sqlalchemy.orm import Session

from app.database import get_db
from app.auth import richiedi_admin, richiedi_utente_attivo
from app.models.corrispettivi import CorrispettiviDocumento, CorrispettiviImport, CorrispettiviManuale
from app.routers.corrispettivi_shared import (
    STRUTTURE_HOTEL, STRUTTURE_MANUALI, STRUTTURE_ORDINE, NOME_STRUTTURA, CATEGORIE, _to_float,
)

router = APIRouter()


@router.get("/report/giornaliero")
def report_giornaliero(
    data_da: Optional[date] = Query(None),
    data_a: Optional[date] = Query(None),
    struttura_code: Optional[str] = Query(None),
    tipo: str = Query('tutti', description="'scontrini'|'fatture'|'tutti'"),
    lordo: bool = Query(True),
    is_test: bool = Query(False),
    db: Session = Depends(get_db),
    _=Depends(richiedi_utente_attivo),
):
    """Aggregato giornaliero per struttura con dettaglio per categoria."""

    # Aliquote per categoria (per conversione lordo→netto)
    ALIQ: dict = {
        'arrangiamenti': 10.0,
        'tassa_soggiorno': 0.0,
        'penali': 0.0,
        'shop': 22.0,
        'altro': 10.0,
    }

    def _v(val_lordo: float, cat: str) -> float:
        if lordo:
            return round(val_lordo, 2)
        aliq = ALIQ.get(cat, 10.0)
        return round(val_lordo / (1 + aliq / 100), 2) if aliq else round(val_lordo, 2)

    # Filtra tipi documenti da includere
    tipi_filtro: List[str] = []
    if tipo == 'scontrini':
        tipi_filtro = ['scontrino']
    elif tipo == 'fatture':
        tipi_filtro = ['fattura']
    else:
        tipi_filtro = ['scontrino', 'fattura']

    # Aggrega da corrispettivi_documenti
    q = (
        db.query(
            CorrispettiviDocumento.data_documento,
            CorrispettiviDocumento.struttura_code,
            CorrispettiviDocumento.tipo,
            CorrispettiviDocumento.categoria,
            func.sum(CorrispettiviDocumento.totale_lordo).label('totale_lordo'),
            func.sum(CorrispettiviDocumento.iva).label('sum_iva'),
            # Somma tassa_soggiorno (NULL se colonna assente nel file sorgente)
            func.sum(CorrispettiviDocumento.tassa_soggiorno).label('sum_ts'),
            func.count().label('n_tot'),
            func.count().filter(CorrispettiviDocumento.annullato == True).label('n_annullati'),
        )
        .filter(
            CorrispettiviDocumento.tipo.in_(tipi_filtro),
            CorrispettiviDocumento.is_test == is_test,
        )
    )
    if data_da:
        q = q.filter(CorrispettiviDocumento.data_documento >= data_da)
    if data_a:
        q = q.filter(CorrispettiviDocumento.data_documento <= data_a)
    if struttura_code:
        q = q.filter(CorrispettiviDocumento.struttura_code == struttura_code.upper())

    righe_doc = q.group_by(
        CorrispettiviDocumento.data_documento,
        CorrispettiviDocumento.struttura_code,
        CorrispettiviDocumento.tipo,
        CorrispettiviDocumento.categoria,
    ).all()

    # Aggrega manuali
    q_man = db.query(CorrispettiviManuale).filter(CorrispettiviManuale.is_test == is_test)
    if data_da:
        q_man = q_man.filter(CorrispettiviManuale.data_giorno >= data_da)
    if data_a:
        q_man = q_man.filter(CorrispettiviManuale.data_giorno <= data_a)
    if struttura_code and struttura_code.upper() not in STRUTTURE_MANUALI:
        manuali_rows = []
    else:
        if struttura_code:
            q_man = q_man.filter(CorrispettiviManuale.struttura_code == struttura_code.upper())
        manuali_rows = q_man.all()

    # Struttura dati: date → struttura → tipo → categoria → totale_lordo
    per_data: dict = {}

    for r in righe_doc:
        dkey = r.data_documento.isoformat()
        if dkey not in per_data:
            per_data[dkey] = {}
        sc = r.struttura_code
        if sc not in per_data[dkey]:
            per_data[dkey][sc] = {
                'scontrino': {'cat': {c: 0.0 for c in CATEGORIE}, 'n': 0, 'n_ann': 0},
                'fattura':   {'cat': {c: 0.0 for c in CATEGORIE}, 'n': 0, 'n_ann': 0},
            }
        t = r.tipo
        cat = r.categoria or 'altro'
        non_ann = int(r.n_tot or 0) - int(r.n_annullati or 0)
        lordo_tot = _to_float(r.totale_lordo)
        sum_iva = _to_float(r.sum_iva)
        # sum_ts: None = colonna tassa_soggiorno assente nel file (vecchio formato)
        #         float = valore esatto disponibile (nuovo formato esteso)
        sum_ts = None if r.sum_ts is None else float(r.sum_ts)

        # Disaggregazione mix arrangiamenti + tassa soggiorno per documenti con cat in
        # ('arrangiamenti', 'altro'): aliquota effettiva < 10% indica presenza di TS esente
        if cat in ('arrangiamenti', 'altro'):
            if sum_ts is not None:
                # Nuovo formato: valore esatto dalla colonna 'Tassa di soggiorno'
                lordo_ts = round(sum_ts, 2)
                lordo_arr = round(max(0.0, lordo_tot - lordo_ts), 2)
                per_data[dkey][sc][t]['cat']['arrangiamenti'] += lordo_arr
                if lordo_ts > 0:
                    per_data[dkey][sc][t]['cat']['tassa_soggiorno'] += lordo_ts
            elif sum_iva > 0:
                # Fallback formato base: inferenza da IVA (imp=iva*10, lordo_arr=iva*11)
                lordo_arr = round(sum_iva * 11, 2)
                lordo_ts = round(max(0.0, lordo_tot - lordo_arr), 2)
                per_data[dkey][sc][t]['cat']['arrangiamenti'] += lordo_arr
                if lordo_ts > 0:
                    per_data[dkey][sc][t]['cat']['tassa_soggiorno'] += lordo_ts
            else:
                per_data[dkey][sc][t]['cat'][cat] += lordo_tot
        else:
            per_data[dkey][sc][t]['cat'][cat] += lordo_tot

        per_data[dkey][sc][t]['n'] += non_ann
        per_data[dkey][sc][t]['n_ann'] += int(r.n_annullati or 0)

    # Integra manuali
    per_data_manuali: dict = {}
    for m in manuali_rows:
        dkey = m.data_giorno.isoformat()
        if dkey not in per_data_manuali:
            per_data_manuali[dkey] = {}
        per_data_manuali[dkey][m.struttura_code] = _to_float(m.arrangiamenti_lordo)

    # Raccoglie tutte le date presenti
    tutte_date = sorted(set(list(per_data.keys()) + list(per_data_manuali.keys())))

    risultato = []
    for dkey in tutte_date:
        strutture_giorno = []
        totale_giorno = 0.0

        strutture_in_giorno: Set[str] = set()
        if dkey in per_data:
            strutture_in_giorno.update(per_data[dkey].keys())
        if dkey in per_data_manuali:
            strutture_in_giorno.update(per_data_manuali[dkey].keys())

        for sc in STRUTTURE_ORDINE:
            if sc not in strutture_in_giorno:
                continue

            dati_sc = per_data.get(dkey, {}).get(sc, {
                'scontrino': {'cat': {c: 0.0 for c in CATEGORIE}, 'n': 0, 'n_ann': 0},
                'fattura':   {'cat': {c: 0.0 for c in CATEGORIE}, 'n': 0, 'n_ann': 0},
            })

            def _cat_dict(t: str) -> dict:
                cats = dati_sc[t]['cat']
                return {
                    cat: _v(cats.get(cat, 0.0), cat)
                    for cat in CATEGORIE
                }

            sc_cats = _cat_dict('scontrino')
            f_cats  = _cat_dict('fattura')

            sc_tot = round(sum(sc_cats.values()), 2)
            f_tot  = round(sum(f_cats.values()), 2)

            man_lordo = _to_float(per_data_manuali.get(dkey, {}).get(sc, 0))
            man = _v(man_lordo, 'arrangiamenti') if sc in STRUTTURE_MANUALI else 0.0

            totale_struttura = sc_tot + f_tot + man
            totale_giorno += totale_struttura

            strutture_giorno.append({
                'struttura_code': sc,
                'struttura_nome': NOME_STRUTTURA.get(sc, sc),
                'scontrini': {**sc_cats, 'totale': sc_tot},
                'fatture':   {**f_cats,  'totale': f_tot},
                'manuale': man,
                'totale': round(totale_struttura, 2),
                'n_scontrini': dati_sc['scontrino']['n'],
                'n_fatture':   dati_sc['fattura']['n'],
                'n_annullati': dati_sc['scontrino']['n_ann'] + dati_sc['fattura']['n_ann'],
            })

        risultato.append({
            'data': dkey,
            'strutture': strutture_giorno,
            'totale_giorno': round(totale_giorno, 2),
        })

    return risultato


@router.get("/report/mensile")
def report_mensile(
    mese: int = Query(..., ge=1, le=12),
    anno: int = Query(..., ge=2020, le=2030),
    struttura_code: Optional[str] = Query(None),
    tipo: str = Query('tutti'),
    lordo: bool = Query(True),
    is_test: bool = Query(False),
    db: Session = Depends(get_db),
    _=Depends(richiedi_utente_attivo),
):
    import calendar
    primo = date(anno, mese, 1)
    ultimo = date(anno, mese, calendar.monthrange(anno, mese)[1])

    risposta_gg = report_giornaliero(
        data_da=primo, data_a=ultimo,
        struttura_code=struttura_code,
        tipo=tipo, lordo=lordo, is_test=is_test,
        db=db, _=None,
    )

    aggregati: dict = {}
    for giorno in risposta_gg:
        for s in giorno['strutture']:
            sc = s['struttura_code']
            if sc not in aggregati:
                aggregati[sc] = {
                    'struttura_code': sc,
                    'struttura_nome': s['struttura_nome'],
                    'scontrini': {c: 0.0 for c in CATEGORIE + ['totale']},
                    'fatture':   {c: 0.0 for c in CATEGORIE + ['totale']},
                    'manuale': 0.0, 'totale': 0.0,
                    'n_scontrini': 0, 'n_fatture': 0, 'n_annullati': 0,
                }
            agg = aggregati[sc]
            for cat in CATEGORIE + ['totale']:
                agg['scontrini'][cat] += s['scontrini'].get(cat, 0)
                agg['fatture'][cat]   += s['fatture'].get(cat, 0)
            agg['manuale']      += s['manuale']
            agg['totale']       += s['totale']
            agg['n_scontrini']  += s['n_scontrini']
            agg['n_fatture']    += s['n_fatture']
            agg['n_annullati']  += s['n_annullati']

    strutture_out = [aggregati[sc] for sc in STRUTTURE_ORDINE if sc in aggregati]
    for sc in sorted(aggregati):
        if sc not in STRUTTURE_ORDINE and sc not in [s['struttura_code'] for s in strutture_out]:
            strutture_out.append(aggregati[sc])

    totale_generale = sum(s['totale'] for s in strutture_out)
    return {
        'mese': mese,
        'anno': anno,
        'strutture': strutture_out,
        'totale_generale': round(totale_generale, 2),
    }


@router.get("/check")
def check_totali(
    data_da: Optional[date] = Query(None),
    data_a: Optional[date] = Query(None),
    lordo: bool = Query(True),
    is_test: bool = Query(False),
    db: Session = Depends(get_db),
    _=Depends(richiedi_utente_attivo),
):
    """Totali aggregati per struttura separando hotel e ristoranti."""
    risposta_gg = report_giornaliero(
        data_da=data_da, data_a=data_a,
        struttura_code=None, tipo='tutti',
        lordo=lordo, is_test=is_test,
        db=db, _=None,
    )

    per_struttura: dict = {s: 0.0 for s in STRUTTURE_ORDINE}
    for giorno in risposta_gg:
        for s in giorno['strutture']:
            sc = s['struttura_code']
            if sc in per_struttura:
                per_struttura[sc] += s['totale']

    totale_hotel = sum(per_struttura.get(s, 0) for s in STRUTTURE_HOTEL)
    totale_ristoranti = sum(per_struttura.get(s, 0) for s in STRUTTURE_MANUALI)
    totale_generale = totale_hotel + totale_ristoranti

    return {
        **{sc: round(per_struttura.get(sc, 0), 2) for sc in STRUTTURE_ORDINE},
        'totale_hotel': round(totale_hotel, 2),
        'totale_ristoranti': round(totale_ristoranti, 2),
        'totale_generale': round(totale_generale, 2),
        'label_hotel': ' + '.join(STRUTTURE_HOTEL),
        'label_ristoranti': ' + '.join(STRUTTURE_MANUALI),
    }


@router.get("/report/fatturati")
def report_fatturati(
    anno: int = Query(..., ge=2020, le=2030),
    lordo: bool = Query(True),
    is_test: bool = Query(False),
    db: Session = Depends(get_db),
    _=Depends(richiedi_utente_attivo),
):
    """Riepilogo fatturati per mese e struttura nell'anno.

    Aggrega scontrini e fatture da corrispettivi_documenti
    (annullato=false) più corrispettivi_manuali per MMS/BON.
    Tassa soggiorno esclusa dal totale (transito verso il Comune).
    Se lordo=false applica divisori per categoria.
    """
    from app.utils.locale_it import MESI_IT

    ALIQ: dict = {
        'arrangiamenti': 10.0,
        'tassa_soggiorno': 0.0,
        'penali': 0.0,
        'shop': 22.0,
        'altro': 10.0,
    }

    def _v(val_lordo: float, cat: str) -> float:
        if lordo:
            return round(val_lordo, 2)
        aliq = ALIQ.get(cat, 10.0)
        return round(val_lordo / (1 + aliq / 100), 2) if aliq else round(val_lordo, 2)

    # ── Query documenti aggregata per (mese, struttura, categoria) ────────────
    mese_col = func.extract('month', CorrispettiviDocumento.data_documento).label('mese')
    righe_doc = (
        db.query(
            mese_col,
            CorrispettiviDocumento.struttura_code,
            CorrispettiviDocumento.categoria,
            func.sum(CorrispettiviDocumento.totale_lordo).label('totale_lordo'),
            func.sum(CorrispettiviDocumento.iva).label('sum_iva'),
            func.sum(CorrispettiviDocumento.tassa_soggiorno).label('sum_ts'),
        )
        .filter(
            func.extract('year', CorrispettiviDocumento.data_documento) == anno,
            CorrispettiviDocumento.tipo.in_(['scontrino', 'fattura']),
            CorrispettiviDocumento.annullato == False,
            CorrispettiviDocumento.is_test == is_test,
        )
        .group_by(mese_col, CorrispettiviDocumento.struttura_code, CorrispettiviDocumento.categoria)
        .all()
    )

    # ── Query manuali aggregata per (mese, struttura) ─────────────────────────
    mese_man_col = func.extract('month', CorrispettiviManuale.data_giorno).label('mese')
    righe_man = (
        db.query(
            mese_man_col,
            CorrispettiviManuale.struttura_code,
            func.sum(CorrispettiviManuale.arrangiamenti_lordo).label('arrangiamenti_lordo'),
        )
        .filter(
            func.extract('year', CorrispettiviManuale.data_giorno) == anno,
            CorrispettiviManuale.is_test == is_test,
        )
        .group_by(mese_man_col, CorrispettiviManuale.struttura_code)
        .all()
    )

    # ── Accumulo per (mese, struttura) ────────────────────────────────────────
    # struttura → {arrangiamenti, tassa_soggiorno, penali, shop, altro}
    def _empty_cats() -> dict:
        return {c: 0.0 for c in CATEGORIE}

    per_mese: dict = {}  # mese(int) → struttura → cats

    for r in righe_doc:
        m = int(r.mese)
        sc = r.struttura_code
        cat = r.categoria or 'altro'
        lordo_tot = _to_float(r.totale_lordo)
        sum_iva = _to_float(r.sum_iva)
        sum_ts = None if r.sum_ts is None else float(r.sum_ts)

        if m not in per_mese:
            per_mese[m] = {}
        if sc not in per_mese[m]:
            per_mese[m][sc] = _empty_cats()

        # Stessa disaggregazione usata nel report giornaliero
        if cat in ('arrangiamenti', 'altro'):
            if sum_ts is not None:
                lordo_ts = round(sum_ts, 2)
                lordo_arr = round(max(0.0, lordo_tot - lordo_ts), 2)
                per_mese[m][sc]['arrangiamenti'] += lordo_arr
                if lordo_ts > 0:
                    per_mese[m][sc]['tassa_soggiorno'] += lordo_ts
            elif sum_iva > 0:
                lordo_arr = round(sum_iva * 11, 2)
                lordo_ts = round(max(0.0, lordo_tot - lordo_arr), 2)
                per_mese[m][sc]['arrangiamenti'] += lordo_arr
                if lordo_ts > 0:
                    per_mese[m][sc]['tassa_soggiorno'] += lordo_ts
            else:
                per_mese[m][sc][cat] += lordo_tot
        else:
            per_mese[m][sc][cat] += lordo_tot

    for r in righe_man:
        m = int(r.mese)
        sc = r.struttura_code
        arr = _to_float(r.arrangiamenti_lordo)
        if m not in per_mese:
            per_mese[m] = {}
        if sc not in per_mese[m]:
            per_mese[m][sc] = _empty_cats()
        per_mese[m][sc]['arrangiamenti'] += arr

    if not per_mese:
        return {
            'anno': anno,
            'strutture': [],
            'mesi': [],
            'totale_anno': {
                'per_struttura': {},
                'totale_hotel': 0.0,
                'totale_ristoranti': 0.0,
                'totale_generale': 0.0,
            },
        }

    # Strutture presenti nell'anno, ordinate per STRUTTURE_ORDINE
    strutture_presenti = [
        s for s in STRUTTURE_ORDINE
        if any(s in per_mese[m] for m in per_mese)
    ]

    def _struttura_output(cats_lordo: dict) -> dict:
        """Converte i valori lordi in output (lordo o netto), calcola totale senza TS."""
        out = {c: _v(cats_lordo.get(c, 0.0), c) for c in CATEGORIE}
        # totale esclude tassa_soggiorno e altro
        out['totale'] = round(
            out['arrangiamenti'] + out['penali'] + out['shop'] + out['altro'], 2
        )
        return out

    # ── Totale anno per struttura ─────────────────────────────────────────────
    totale_anno_cats: dict = {s: _empty_cats() for s in strutture_presenti}
    for m, per_sc in per_mese.items():
        for sc, cats in per_sc.items():
            if sc in totale_anno_cats:
                for c in CATEGORIE:
                    totale_anno_cats[sc][c] += cats.get(c, 0.0)

    totale_anno_per_struttura = {
        sc: _struttura_output(totale_anno_cats[sc])
        for sc in strutture_presenti
    }
    tot_anno_hotel = round(sum(
        totale_anno_per_struttura[s]['totale']
        for s in strutture_presenti if s in STRUTTURE_HOTEL
    ), 2)
    tot_anno_rist = round(sum(
        totale_anno_per_struttura[s]['totale']
        for s in strutture_presenti if s in STRUTTURE_MANUALI
    ), 2)

    # ── Costruzione risposta mesi ─────────────────────────────────────────────
    mesi_output = []
    for m in sorted(per_mese.keys()):
        per_sc = per_mese[m]
        per_struttura_out = {}
        for sc in strutture_presenti:
            cats = per_sc.get(sc, _empty_cats())
            per_struttura_out[sc] = _struttura_output(cats)

        tot_hotel = round(sum(
            per_struttura_out[s]['totale']
            for s in strutture_presenti if s in STRUTTURE_HOTEL
        ), 2)
        tot_rist = round(sum(
            per_struttura_out[s]['totale']
            for s in strutture_presenti if s in STRUTTURE_MANUALI
        ), 2)

        mesi_output.append({
            'mese': m,
            'nome_mese': MESI_IT[m - 1],
            'per_struttura': per_struttura_out,
            'totale_hotel': tot_hotel,
            'totale_ristoranti': tot_rist,
            'totale_generale': round(tot_hotel + tot_rist, 2),
        })

    return {
        'anno': anno,
        'strutture': strutture_presenti,
        'mesi': mesi_output,
        'totale_anno': {
            'per_struttura': totale_anno_per_struttura,
            'totale_hotel': tot_anno_hotel,
            'totale_ristoranti': tot_anno_rist,
            'totale_generale': round(tot_anno_hotel + tot_anno_rist, 2),
        },
    }


@router.get("/report/pagamenti")
def report_pagamenti(
    anno: int = Query(..., ge=2020, le=2030),
    is_test: bool = Query(False),
    db: Session = Depends(get_db),
    _=Depends(richiedi_utente_attivo),
):
    """Riepilogo fatturati per tipo di pagamento e mese nell'anno.

    Normalizza il testo grezzo della colonna Pagamenti (es. 'Contante 8,00 € /')
    estraendo solo il nome del tipo.  Restituisce sempre valori lordi.
    """
    from app.utils.locale_it import MESI_IT

    TIPI_NOTI = [
        'Bonifico/Vaglia', 'XPAY-Nexi', 'Carta Credito',
        'Bancomat', 'Bonifico', 'Contante', 'Satispay', 'Assegno', 'xpay',
    ]

    def _normalizza(raw: str) -> str:
        if not raw or not raw.strip():
            return 'Non specificato'
        raw_l = raw.strip().lower()
        for tipo in TIPI_NOTI:
            if raw_l.startswith(tipo.lower()):
                return tipo
        return raw.strip()

    mese_col = func.extract('month', CorrispettiviDocumento.data_documento).label('mese')
    # Ogni documento si divide in tre componenti:
    #   pagato (tipo_pagamento) = totale_lordo - sospeso - deposito - tassa_soggiorno
    #   caparra (deposito)      = deposito
    #   sospeso                 = sospeso
    # Totale = pagato + caparra + sospeso = totale_lordo - TS  →  pareggia con report_fatturati.
    from sqlalchemy import func as sqlfunc
    pagato_col  = func.sum(
        CorrispettiviDocumento.totale_lordo
        - CorrispettiviDocumento.sospeso
        - CorrispettiviDocumento.deposito
        - sqlfunc.coalesce(CorrispettiviDocumento.tassa_soggiorno, 0)
    ).label('totale_pagato')
    sospeso_col = func.sum(CorrispettiviDocumento.sospeso).label('totale_sospeso')
    deposito_col = func.sum(CorrispettiviDocumento.deposito).label('totale_deposito')

    righe = (
        db.query(mese_col, CorrispettiviDocumento.tipo_pagamento,
                 pagato_col, sospeso_col, deposito_col)
        .filter(
            func.extract('year', CorrispettiviDocumento.data_documento) == anno,
            CorrispettiviDocumento.tipo.in_(['scontrino', 'fattura']),
            CorrispettiviDocumento.annullato == False,
            CorrispettiviDocumento.is_test == is_test,
        )
        .group_by(mese_col, CorrispettiviDocumento.tipo_pagamento)
        .all()
    )

    # Aggiungi i corrispettivi manuali MMS/BON per mese
    mese_man_col = func.extract('month', CorrispettiviManuale.data_giorno).label('mese')
    righe_man = (
        db.query(mese_man_col, func.sum(CorrispettiviManuale.arrangiamenti_lordo).label('totale'))
        .filter(
            func.extract('year', CorrispettiviManuale.data_giorno) == anno,
            CorrispettiviManuale.is_test == is_test,
        )
        .group_by(mese_man_col)
        .all()
    )

    if not righe and not righe_man:
        return {'anno': anno, 'mesi': [], 'tipi': [], 'per_tipo': {},
                'totale_mese': {}, 'totale_anno': {}}

    per_tipo: dict = {}
    mesi_set: set = set()

    def _acc(key: str, mese: int, val: float):
        if val == 0:
            return
        if key not in per_tipo:
            per_tipo[key] = {}
        per_tipo[key][mese] = round(per_tipo[key].get(mese, 0.0) + val, 2)

    for r in righe:
        m = int(r.mese)
        mesi_set.add(m)

        val_pag  = round(float(r.totale_pagato  or 0), 2)
        val_sosp = round(float(r.totale_sospeso or 0), 2)
        val_dep  = round(float(r.totale_deposito or 0), 2)

        if val_pag != 0:
            _acc(_normalizza(str(r.tipo_pagamento or '')), m, val_pag)
        _acc('Caparra', m, val_dep)
        _acc('Sospeso', m, val_sosp)

    for r in righe_man:
        m = int(r.mese)
        mesi_set.add(m)
        _acc('MMS / BON (manuale)', m, round(float(r.totale or 0), 2))

    mesi_ordinati = sorted(mesi_set)

    # Ordine: tipi noti → altri → Non specificato → MMS/BON → Caparra → Sospeso in fondo
    SPECIALI = ('Non specificato', 'MMS / BON (manuale)', 'Caparra', 'Sospeso')
    tipi_noti_presenti = [t for t in TIPI_NOTI if t in per_tipo]
    tipi_altri = sorted(t for t in per_tipo if t not in TIPI_NOTI and t not in SPECIALI)
    tipi_ns    = ['Non specificato']    if 'Non specificato'    in per_tipo else []
    tipi_man   = ['MMS / BON (manuale)'] if 'MMS / BON (manuale)' in per_tipo else []
    tipi_cap   = ['Caparra']            if 'Caparra'            in per_tipo else []
    tipi_sosp  = ['Sospeso']            if 'Sospeso'            in per_tipo else []
    tipi_ordinati = tipi_noti_presenti + tipi_altri + tipi_ns + tipi_man + tipi_cap + tipi_sosp

    totale_mese = {
        m: round(sum(per_tipo[t].get(m, 0.0) for t in tipi_ordinati), 2)
        for m in mesi_ordinati
    }
    totale_anno = {t: round(sum(per_tipo[t].values()), 2) for t in tipi_ordinati}

    return {
        'anno': anno,
        'mesi': [{'mese': m, 'nome_mese': MESI_IT[m - 1]} for m in mesi_ordinati],
        'tipi': tipi_ordinati,
        'per_tipo': {
            t: {str(m): per_tipo[t].get(m, 0.0) for m in mesi_ordinati}
            for t in tipi_ordinati
        },
        'totale_mese': {str(m): totale_mese[m] for m in mesi_ordinati},
        'totale_anno': totale_anno,
    }


@router.get("/admin/test-stats")
def test_stats(
    db: Session = Depends(get_db),
    _=Depends(richiedi_admin),
):
    n_doc = db.query(func.count(CorrispettiviDocumento.id)).filter(
        CorrispettiviDocumento.is_test == True).scalar() or 0
    n_imp = db.query(func.count(CorrispettiviImport.id)).filter(
        CorrispettiviImport.is_test == True).scalar() or 0
    n_man = db.query(func.count(CorrispettiviManuale.id)).filter(
        CorrispettiviManuale.is_test == True).scalar() or 0
    return {
        'imports': n_imp,
        'documenti': n_doc,
        'manuali': n_man,
        'totale': n_imp + n_doc + n_man,
    }


@router.delete("/admin/test-data")
def elimina_test_data(
    conferma: bool = Query(False),
    db: Session = Depends(get_db),
    _=Depends(richiedi_admin),
):
    if not conferma:
        raise HTTPException(status_code=400,
                             detail="Aggiungere ?conferma=true per confermare")

    db.query(CorrispettiviDocumento).filter(
        CorrispettiviDocumento.is_test == True).delete(synchronize_session=False)
    db.query(CorrispettiviManuale).filter(
        CorrispettiviManuale.is_test == True).delete(synchronize_session=False)
    db.query(CorrispettiviImport).filter(
        CorrispettiviImport.is_test == True).delete(synchronize_session=False)

    db.commit()
    return {'eliminati': True}


@router.get("/export/fatturati")
def export_fatturati(
    anno: int = Query(..., ge=2020, le=2030),
    lordo: bool = Query(True),
    is_test: bool = Query(False),
    db: Session = Depends(get_db),
    _=Depends(richiedi_utente_attivo),
):
    """Esporta il riepilogo fatturati in Excel (3 fogli: Corrispettivi, Tassa Soggiorno, Controllo)."""
    import io
    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from app.utils.locale_it import MESI_IT

    # Riusa la logica di report_fatturati
    dati = report_fatturati(anno=anno, lordo=lordo, is_test=is_test, db=db, _=None)
    strutture = dati['strutture']
    mesi = dati['mesi']
    tot_anno = dati['totale_anno']

    strutture_hotel = [s for s in strutture if s in STRUTTURE_HOTEL]

    HDR_FILL = PatternFill('solid', fgColor='1E3A5F')
    HDR_FONT = Font(bold=True, color='FFFFFF', size=9)
    TOT_FILL = PatternFill('solid', fgColor='1E3A5F')
    TOT_FONT = Font(bold=True, color='FFFFFF', size=9)
    ALT_FILL = PatternFill('solid', fgColor='F8FAFC')
    TEAL_FILL = PatternFill('solid', fgColor='0F766E')
    NUM_FMT = '#,##0.00 "€"'

    wb = Workbook()

    def _hdr(ws, row, cols):
        for c, val in enumerate(cols, 1):
            cell = ws.cell(row=row, column=c, value=val)
            cell.font = HDR_FONT
            cell.fill = HDR_FILL
            cell.alignment = Alignment(horizontal='center')

    def _tot_row(ws, row, vals):
        for c, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=c, value=val)
            cell.font = TOT_FONT
            cell.fill = TOT_FILL
            cell.alignment = Alignment(horizontal='right' if c > 1 else 'left')
            if isinstance(val, float):
                cell.number_format = NUM_FMT

    def _data_row(ws, row, vals, alt=False):
        fill = ALT_FILL if alt else None
        for c, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=c, value=val)
            cell.alignment = Alignment(horizontal='right' if c > 1 else 'left')
            if fill:
                cell.fill = fill
            if isinstance(val, float):
                cell.number_format = NUM_FMT

    def _autowidth(ws):
        for col in ws.columns:
            max_len = max((len(str(cell.value or '')) for cell in col), default=8)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 22)

    # ── Foglio 1: Totale Corrispettivi ────────────────────────────────────────
    ws1 = wb.active
    ws1.title = 'Corrispettivi'
    hdrs = ['Mese'] + strutture + ['Tot. Hotel', 'Tot. Rist.', 'TOTALE']
    _hdr(ws1, 1, hdrs)
    for i, m in enumerate(mesi):
        row = i + 2
        vals = [m['nome_mese']]
        for s in strutture:
            vals.append(m['per_struttura'].get(s, {}).get('totale') or 0.0)
        vals += [m.get('totale_hotel') or 0.0,
                 m.get('totale_ristoranti') or 0.0,
                 m.get('totale_generale') or 0.0]
        _data_row(ws1, row, vals, alt=i % 2 == 1)
    tot_vals = ['TOTALE ANNO']
    for s in strutture:
        tot_vals.append(tot_anno['per_struttura'].get(s, {}).get('totale') or 0.0)
    tot_vals += [tot_anno.get('totale_hotel') or 0.0,
                 tot_anno.get('totale_ristoranti') or 0.0,
                 tot_anno.get('totale_generale') or 0.0]
    _tot_row(ws1, len(mesi) + 2, tot_vals)
    _autowidth(ws1)

    # ── Foglio 2: Tassa di Soggiorno ──────────────────────────────────────────
    ws2 = wb.create_sheet('Tassa di Soggiorno')
    hdrs2 = ['Mese'] + strutture_hotel + ['Tot. Hotel']
    _hdr(ws2, 1, hdrs2)
    for i, m in enumerate(mesi):
        row = i + 2
        vals = [m['nome_mese']]
        tot_ts = 0.0
        for s in strutture_hotel:
            ts = m['per_struttura'].get(s, {}).get('tassa_soggiorno') or 0.0
            vals.append(ts)
            tot_ts += ts
        vals.append(tot_ts)
        _data_row(ws2, row, vals, alt=i % 2 == 1)
    tot_vals2 = ['TOTALE ANNO']
    tot_ts_anno = 0.0
    for s in strutture_hotel:
        ts = tot_anno['per_struttura'].get(s, {}).get('tassa_soggiorno') or 0.0
        tot_vals2.append(ts)
        tot_ts_anno += ts
    tot_vals2.append(tot_ts_anno)
    _tot_row(ws2, len(mesi) + 2, tot_vals2)
    _autowidth(ws2)

    # ── Foglio 3: Riepilogo di controllo ─────────────────────────────────────
    ws3 = wb.create_sheet('Riepilogo Controllo')
    hdrs3 = ['Mese', 'Corrispettivo', '+ Tassa Soggiorno', '= Totale Lordo']

    for c, val in enumerate(hdrs3, 1):
        cell = ws3.cell(row=1, column=c, value=val)
        cell.font = HDR_FONT
        cell.fill = TEAL_FILL if c in (2, 3) else HDR_FILL
        cell.alignment = Alignment(horizontal='center')

    ts_per_mese = {
        m['mese']: sum(
            m['per_struttura'].get(s, {}).get('tassa_soggiorno') or 0.0
            for s in strutture
        )
        for m in mesi
    }
    for i, m in enumerate(mesi):
        row = i + 2
        corr = m.get('totale_generale') or 0.0
        ts = ts_per_mese[m['mese']]
        _data_row(ws3, row, [m['nome_mese'], corr, ts, corr + ts], alt=i % 2 == 1)

    ts_anno = sum(
        tot_anno['per_struttura'].get(s, {}).get('tassa_soggiorno') or 0.0
        for s in strutture
    )
    corr_anno = tot_anno.get('totale_generale') or 0.0
    _tot_row(ws3, len(mesi) + 2, ['TOTALE ANNO', corr_anno, ts_anno, corr_anno + ts_anno])
    _autowidth(ws3)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f'corrispettivi_fatturati_{anno}.xlsx'
    return StreamingResponse(
        buf,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'},
    )
