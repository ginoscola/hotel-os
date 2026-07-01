"""Parser Excel per il modulo Corrispettivi v4.

Supporta due formati di export da Welcome PMS:

Formato base (listaConti.xlsx, 18 colonne):
  Data | Numero | Suffisso | Intestazione | Totale | Incassato | Deposito |
  Sospeso | Abbuono | Camere | Codice prenotazione | Pagamenti |
  Conto anticipato | Acconto | Imponibile | Iva | Annullato | Note

Formato esteso (36 colonne, include Tassa di soggiorno):
  Data | Sigla | Numero | Suffisso | Intestazione | Totale | Incassato | Deposito |
  Sospeso | Abbuono | Tassa di soggiorno | Numero Scontrino | Camere | Pagamenti |
  Codice prenotazione | Arrivo | Partenza | Ubicazione Istat | Voucher |
  Imponibile | Iva | Note | Nome File | Stato FE | Modalità | Importo bollo |
  Tipo documento | Numero documento | Nazione | Ora Stampa | Contabilizzato Mexal |
  Data annullamento | Causale cancellazione | Maschera conto | Data creazione | Utente creazione

Nel formato esteso:
  - Tassa di soggiorno: valore esatto per ogni documento
  - Data annullamento: sostituisce la colonna Annullato (documento annullato se non vuota)

Restituisce tutti i documenti in lista unificata con campo `tipo`:
  'scontrino' — SC/SCA inclusi nel registro
  'fattura'   — F incluse nel registro
  'escluso'   — CP/FD/altri esclusi fiscalmente (salvati per audit)

Le righe con struttura non determinabile finiscono in `righe_non_salvabili`.
"""

import re
from dataclasses import dataclass, field
from datetime import date, datetime
from typing import Dict, List, Optional, Set

import openpyxl


# ── Costanti ─────────────────────────────────────────────────────────────────

PREFISSO_A_STRUTTURA: Dict[str, str] = {
    'D': 'DPH',
    'C': 'CLB',
    'I': 'INT',
}

CAMERA_SPECIALE_A_STRUTTURA: Dict[str, str] = {
    'AIRE': 'DPH',
    'AGUA': 'DPH',
    'TIERRA': 'DPH',
    'FUEGO': 'DPH',
}

TIPI_SCONTRINO: Set[str] = {'SC', 'SCA'}
TIPI_FATTURA: Set[str] = {'F'}

TOLLERANZA_ALIQUOTA = 0.5

# Tipi pagamento noti, dal più specifico al meno specifico (per il matching startsWith).
# Usati per pulire la stringa grezza della colonna "Pagamenti" che include l'importo.
_TIPI_PAG_NOTI = [
    'Bonifico/Vaglia', 'XPAY-Nexi', 'Carta Credito',
    'Bancomat', 'Bonifico', 'Contante', 'Satispay', 'Assegno', 'xpay',
]

def _estrai_tipo_pagamento(raw: str) -> str:
    """Da 'Contante 8,00 € / Carta Credito 12,00 € /' estrae il primo tipo noto.
    Se non trova corrispondenza restituisce la stringa originale."""
    if not raw:
        return raw
    raw_l = raw.lower()
    for tipo in _TIPI_PAG_NOTI:
        if raw_l.startswith(tipo.lower()):
            return tipo
    return raw


# ── Strutture dati ────────────────────────────────────────────────────────────

@dataclass
class RigaDocumento:
    data_documento: date
    numero: Optional[int]
    suffisso: str
    tipo: str           # 'scontrino' | 'fattura' | 'escluso'
    struttura_code: str
    intestazione: str
    camera: str
    totale_lordo: float
    incassato: float
    deposito: float
    sospeso: float
    abbuono: float
    imponibile: float
    iva: float
    aliquota_pct: float
    categoria: Optional[str]   # None per tipo='escluso'
    codice_prenotazione: str
    tipo_pagamento: str
    conto_anticipato: bool
    acconto: bool
    annullato: bool
    note: str
    ospiti: str
    motivo_esclusione: Optional[str] = None   # solo per tipo='escluso'
    # Campi formato esteso Welcome PMS (None se file in formato base 18 colonne)
    tassa_soggiorno: Optional[float] = None
    sigla: Optional[str] = None
    numero_scontrino: Optional[str] = None
    arrivo: Optional[date] = None
    partenza: Optional[date] = None
    ubicazione_istat: Optional[str] = None
    voucher: Optional[str] = None
    nome_file_pms: Optional[str] = None
    stato_fe: Optional[str] = None
    modalita: Optional[str] = None
    importo_bollo: Optional[float] = None
    tipo_documento_fe: Optional[str] = None
    numero_documento_fe: Optional[str] = None
    nazione: Optional[str] = None
    ora_stampa: Optional[str] = None
    contabilizzato_mexal: Optional[str] = None
    causale_cancellazione: Optional[str] = None
    maschera_conto: Optional[str] = None
    data_creazione_doc: Optional[date] = None
    utente_creazione: Optional[str] = None


@dataclass
class RigaEsclusa:
    """Riga non salvabile: struttura non determinabile."""
    riga_excel: int
    motivo: str
    suffisso: str
    numero: Optional[int]
    data: Optional[date]


@dataclass
class RisultatoParsing:
    documenti: List[RigaDocumento] = field(default_factory=list)
    righe_non_salvabili: List[RigaEsclusa] = field(default_factory=list)
    strutture_trovate: Set[str] = field(default_factory=set)   # solo scontrini/fatture
    data_da: Optional[date] = None
    data_a: Optional[date] = None
    warnings: List[str] = field(default_factory=list)

    # Proprietà di compatibilità
    @property
    def scontrini(self) -> List[RigaDocumento]:
        return [d for d in self.documenti if d.tipo == 'scontrino']

    @property
    def fatture(self) -> List[RigaDocumento]:
        return [d for d in self.documenti if d.tipo == 'fattura']

    @property
    def esclusi(self) -> List[RigaDocumento]:
        return [d for d in self.documenti if d.tipo == 'escluso']


# ── Funzioni di utilità ───────────────────────────────────────────────────────

def _to_float(v) -> float:
    if v is None:
        return 0.0
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def _to_int(v) -> Optional[int]:
    if v is None:
        return None
    try:
        return int(v)
    except (TypeError, ValueError):
        return None


def _to_bool(v) -> bool:
    if v is None:
        return False
    if isinstance(v, bool):
        return v
    try:
        return float(v) != 0
    except (TypeError, ValueError):
        return bool(v)


def _parse_data(v) -> Optional[date]:
    """Accetta datetime Python (da openpyxl) o stringa in vari formati."""
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    # Rimuovi eventuale suffisso con giorno della settimana: "01/06/2026 lun"
    s_base = s.split(' ')[0] if ' ' in s else s
    for fmt in ('%d/%m/%Y', '%d/%m/%y', '%Y-%m-%d'):
        for candidato in (s, s_base):
            try:
                return datetime.strptime(candidato, fmt).date()
            except ValueError:
                continue
    return None


def _struttura_da_suffisso(suffisso: str, camera: str) -> Optional[str]:
    """
    Determina struttura_code dal campo Suffisso (es. 'D-SC' → 'DPH').
    Se il prefisso non è riconosciuto prova a usare il campo Camera.
    """
    if suffisso and '-' in suffisso:
        prefisso = suffisso.split('-')[0].upper()
        if prefisso in PREFISSO_A_STRUTTURA:
            return PREFISSO_A_STRUTTURA[prefisso]

    if camera:
        prima = camera.strip().upper()
        for nome, struttura in CAMERA_SPECIALE_A_STRUTTURA.items():
            if prima.startswith(nome):
                return struttura
        if prima and prima[0] in PREFISSO_A_STRUTTURA:
            return PREFISSO_A_STRUTTURA[prima[0]]

    return None


def _tipo_da_suffisso(suffisso: str) -> str:
    """Estrae il tipo documento dal Suffisso (es. 'D-SC' → 'SC')."""
    if suffisso and '-' in suffisso:
        return suffisso.split('-', 1)[1].upper()
    return (suffisso or '').upper()


def _calcola_aliquota(imponibile: float, iva: float) -> float:
    """Calcola aliquota IVA in % da imponibile e iva."""
    if imponibile and imponibile != 0:
        return round(iva / imponibile * 100, 2)
    return 0.0


def _vicino(val: float, target: float) -> bool:
    return abs(val - target) <= TOLLERANZA_ALIQUOTA


def _determina_categoria(
    aliquota_pct: float,
    imponibile: float,
    causale_cancellazione: Optional[str] = None,
    tassa_soggiorno_col: Optional[float] = None,
    colonna_ts_presente: bool = False,
) -> str:
    """
    Determina la categoria del documento in base all'aliquota IVA.

    arrangiamenti   ≈ 10%  (include mix arrangiamenti+tassa soggiorno: 0 < aliq < 10)
    shop            ≈ 22%
    tassa_soggiorno = 0% con imponibile > 0 E colonna TS > 0 (formato esteso)
                      oppure 0% con imponibile > 0 (formato base, ambiguo)
    penali          = 0% con imponibile = 0
                    — oppure 0% con causale_cancellazione valorizzata
                    — oppure 0% con imponibile > 0 MA colonna TS = 0/assente (formato esteso)

    Meccanismo per distinguere TS da penale (Welcome PMS compila imponibile anche
    per documenti fuori campo IVA):
    - Formato esteso: la colonna "Tassa di soggiorno" è la fonte di verità.
      TS > 0 → tassa_soggiorno; TS = 0/vuoto → penale.
    - causale_cancellazione valorizzata → penale certa (segnale aggiuntivo).
    - Formato base (colonna assente): non distinguibile, assume tassa_soggiorno.

    Nota: fatture con mix arrangiamenti (10%) + tassa soggiorno (0%) hanno aliquota
    effettiva tra 0% e 10%. Vengono categorizzate come 'arrangiamenti' — il report
    disaggrega la componente tassa soggiorno usando iva * 11 come lordo arrangiamenti.
    """
    if _vicino(aliquota_pct, 10.0):
        return 'arrangiamenti'
    if _vicino(aliquota_pct, 22.0):
        return 'shop'
    if _vicino(aliquota_pct, 0.0):
        if causale_cancellazione and causale_cancellazione.strip():
            return 'penali'
        # Usa abs: un annullo (-6€) si categorizza come l'originale (+6€)
        if imponibile and abs(imponibile) > 0:
            if colonna_ts_presente:
                return 'tassa_soggiorno' if tassa_soggiorno_col else 'penali'
            return 'tassa_soggiorno'
        return 'penali'
    # Aliquota tra 0% e 10% (esclusivo): mix arrangiamenti + tassa soggiorno
    if 0 < aliquota_pct < 10.0 - TOLLERANZA_ALIQUOTA:
        return 'arrangiamenti'
    return 'altro'


def _estrai_ospiti(note: str) -> str:
    """Estrae la lista ospiti dal campo Note dopo 'Ospiti: '."""
    if not note:
        return ''
    m = re.search(r'Ospiti:\s*(.+)', str(note), re.IGNORECASE)
    return m.group(1).strip() if m else ''


# ── Funzione principale ───────────────────────────────────────────────────────

def parse_excel(file_path: str) -> RisultatoParsing:
    """
    Analizza il file listaConti.xlsx e restituisce tutti i documenti.

    Returns:
        RisultatoParsing con:
          - documenti: scontrini + fatture + esclusi con struttura nota
          - righe_non_salvabili: righe con struttura non determinabile
          - strutture_trovate: strutture di scontrini e fatture (non esclusi)
    """
    risultato = RisultatoParsing()

    wb = openpyxl.load_workbook(file_path, data_only=True)

    # Cerca il foglio con le colonne obbligatorie; parte dal foglio attivo poi prova gli altri
    obbligatorie = ['Data', 'Numero', 'Suffisso', 'Totale', 'Imponibile', 'Iva']
    fogli_da_provare = [wb.active] + [ws for ws in wb.worksheets if ws is not wb.active]
    ws = None
    intestazioni: dict = {}
    for foglio in fogli_da_provare:
        hdrs = {foglio.cell(1, c).value: c for c in range(1, foglio.max_column + 1)}
        if all(col in hdrs for col in obbligatorie):
            ws = foglio
            intestazioni = hdrs
            break

    if ws is None:
        # Nessun foglio ha le colonne attese — usa il foglio attivo e segnala warning
        ws = wb.active
        intestazioni = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
        nomi_fogli = ', '.join(repr(s.title) for s in wb.worksheets)
        risultato.warnings.append(
            f"Colonne obbligatorie non trovate in nessun foglio ({nomi_fogli}). "
            f"Colonne trovate nel foglio attivo: {list(intestazioni.keys())}"
        )

    def _col(nome: str) -> Optional[int]:
        return intestazioni.get(nome)

    # Formato esteso: colonna TS è fonte di verità per tassa_soggiorno vs penali.
    colonna_ts_presente = 'Tassa di soggiorno' in intestazioni

    # Helper per campi formato esteso: definiti una volta sola, ricevono row_idx esplicitamente.
    # _col è stabile (intestazioni non cambia); ws.cell() è la sola chiamata per-riga.
    def _str_ext(row: int, col: str) -> Optional[str]:
        c = _col(col)
        if c is None:
            return None
        v = ws.cell(row, c).value
        return str(v).strip() if v is not None else None

    def _date_ext(row: int, col: str) -> Optional[date]:
        c = _col(col)
        if c is None:
            return None
        return _parse_data(ws.cell(row, c).value)

    def _float_ext(row: int, col: str) -> Optional[float]:
        c = _col(col)
        if c is None:
            return None
        v = ws.cell(row, c).value
        return _to_float(v) if v is not None else None

    data_da: Optional[date] = None
    data_a:  Optional[date] = None

    for row_idx in range(2, ws.max_row + 1):
        def _val(nome: str):
            c = _col(nome)
            return ws.cell(row_idx, c).value if c else None

        raw_data     = _val('Data')
        raw_suffisso = str(_val('Suffisso') or '').strip()
        raw_numero   = _val('Numero')
        camera       = str(_val('Camere') or '').strip()

        data_doc = _parse_data(raw_data)
        numero   = _to_int(raw_numero)
        tipo_doc = _tipo_da_suffisso(raw_suffisso)

        if data_doc is None:
            if raw_data is not None:
                risultato.warnings.append(
                    f"Riga {row_idx}: data non valida '{raw_data}', riga saltata"
                )
            continue

        struttura = _struttura_da_suffisso(raw_suffisso, camera)

        if struttura is None:
            risultato.warnings.append(
                f"Riga {row_idx}: struttura non determinata per suffisso '{raw_suffisso}', "
                f"camera '{camera}' — riga saltata"
            )
            risultato.righe_non_salvabili.append(RigaEsclusa(
                riga_excel=row_idx,
                motivo=f"Struttura non determinata (suffisso={raw_suffisso!r}, camera={camera!r})",
                suffisso=raw_suffisso,
                numero=numero,
                data=data_doc,
            ))
            continue

        if tipo_doc in TIPI_SCONTRINO:
            tipo_riga  = 'scontrino'
            motivo_esc = None
        elif tipo_doc in TIPI_FATTURA:
            tipo_riga  = 'fattura'
            motivo_esc = None
        else:
            tipo_riga  = 'escluso'
            motivo_esc = f"Tipo '{tipo_doc}' escluso fiscalmente (CP/FD/altro)"

        # Campi finanziari
        totale_lordo = _to_float(_val('Totale'))
        incassato    = _to_float(_val('Incassato'))
        deposito     = _to_float(_val('Deposito'))
        sospeso      = _to_float(_val('Sospeso'))
        abbuono      = _to_float(_val('Abbuono'))
        imponibile   = _to_float(_val('Imponibile'))
        iva          = _to_float(_val('Iva'))
        aliquota_pct = _calcola_aliquota(imponibile, iva)

        # Annullato: colonna 'Annullato' (base) o 'Data annullamento' (esteso)
        annullato = _to_bool(_val('Annullato')) or bool(_val('Data annullamento'))

        tipo_pagamento   = _estrai_tipo_pagamento(str(_val('Pagamenti') or '').strip())
        conto_anticipato = _to_bool(_val('Conto anticipato'))
        acconto          = _to_bool(_val('Acconto'))

        note         = str(_val('Note') or '').strip()
        ospiti       = _estrai_ospiti(note)
        intestazione = str(_val('Intestazione') or '').strip()
        cod_pren     = str(_val('Codice prenotazione') or '').strip()

        # Campi formato esteso
        tassa_soggiorno      = _float_ext(row_idx, 'Tassa di soggiorno')
        sigla                = _str_ext(row_idx, 'Sigla')
        numero_scontrino     = _str_ext(row_idx, 'Numero Scontrino')
        arrivo               = _date_ext(row_idx, 'Arrivo')
        partenza             = _date_ext(row_idx, 'Partenza')
        ubicazione_istat     = _str_ext(row_idx, 'Ubicazione Istat')
        voucher              = _str_ext(row_idx, 'Voucher')
        nome_file_pms        = _str_ext(row_idx, 'Nome File')
        stato_fe             = _str_ext(row_idx, 'Stato FE')
        modalita             = _str_ext(row_idx, 'Modalità')
        importo_bollo        = _float_ext(row_idx, 'Importo bollo')
        tipo_documento_fe    = _str_ext(row_idx, 'Tipo documento')
        numero_documento_fe  = _str_ext(row_idx, 'Numero documento')
        nazione              = _str_ext(row_idx, 'Nazione')
        ora_stampa           = _str_ext(row_idx, 'Ora Stampa')
        contabilizzato_mexal = _str_ext(row_idx, 'Contabilizzato Mexal')
        causale_cancellazione = _str_ext(row_idx, 'Causale cancellazione')
        maschera_conto       = _str_ext(row_idx, 'Maschera conto')
        data_creazione_doc   = _date_ext(row_idx, 'Data creazione')
        utente_creazione     = _str_ext(row_idx, 'Utente creazione')

        categoria = _determina_categoria(
            aliquota_pct, imponibile,
            causale_cancellazione=causale_cancellazione,
            tassa_soggiorno_col=tassa_soggiorno,
            colonna_ts_presente=colonna_ts_presente,
        ) if tipo_riga != 'escluso' else None

        risultato.documenti.append(RigaDocumento(
            data_documento=data_doc,
            numero=numero,
            suffisso=raw_suffisso,
            tipo=tipo_riga,
            struttura_code=struttura,
            intestazione=intestazione,
            camera=camera,
            totale_lordo=totale_lordo,
            incassato=incassato,
            deposito=deposito,
            sospeso=sospeso,
            abbuono=abbuono,
            imponibile=imponibile,
            iva=iva,
            aliquota_pct=aliquota_pct,
            categoria=categoria,
            codice_prenotazione=cod_pren,
            tipo_pagamento=tipo_pagamento,
            conto_anticipato=conto_anticipato,
            acconto=acconto,
            annullato=annullato,
            note=note,
            ospiti=ospiti,
            motivo_esclusione=motivo_esc,
            tassa_soggiorno=tassa_soggiorno,
            sigla=sigla,
            numero_scontrino=numero_scontrino,
            arrivo=arrivo,
            partenza=partenza,
            ubicazione_istat=ubicazione_istat,
            voucher=voucher,
            nome_file_pms=nome_file_pms,
            stato_fe=stato_fe,
            modalita=modalita,
            importo_bollo=importo_bollo,
            tipo_documento_fe=tipo_documento_fe,
            numero_documento_fe=numero_documento_fe,
            nazione=nazione,
            ora_stampa=ora_stampa,
            contabilizzato_mexal=contabilizzato_mexal,
            causale_cancellazione=causale_cancellazione,
            maschera_conto=maschera_conto,
            data_creazione_doc=data_creazione_doc,
            utente_creazione=utente_creazione,
        ))

        if tipo_riga in ('scontrino', 'fattura'):
            risultato.strutture_trovate.add(struttura)
            if data_da is None or data_doc < data_da:
                data_da = data_doc
            if data_a is None or data_doc > data_a:
                data_a = data_doc

    risultato.data_da = data_da
    risultato.data_a  = data_a
    return risultato
